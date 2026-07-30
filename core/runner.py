from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Optional, List, Dict, Any
import hashlib
import json
import csv
import os
import re
import time

from core.paths import stats_path as stats_path_fn
from core.vendor.playlist_scan_safe import ScanCancelled, scan_folder
from core.vendor.repair_playlist_safe_v5 import RepairCancelled, repair_playlist
from openpyxl import load_workbook

ProgressCb = Callable[[int, str], None]


def atomic_write_text(path: Path, text: str, encoding: str = "utf-8") -> None:
    """Replace a text file only after its complete contents are on disk."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(f".{path.name}.{os.getpid()}.{time.time_ns()}.tmp")
    try:
        tmp.write_text(text, encoding=encoding)
        os.replace(tmp, path)
    finally:
        try:
            if tmp.exists():
                tmp.unlink()
        except OSError:
            pass


def normalized_root_path(value: object) -> str:
    """Return a stable Windows path key without touching the filesystem."""
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        return os.path.normcase(
            os.path.abspath(os.path.normpath(text))
        ).casefold()
    except (OSError, TypeError, ValueError):
        return os.path.normcase(os.path.normpath(text)).casefold()


def roots_overlap(left: object, right: object) -> bool:
    """Return True when either Music Root contains the other."""
    left_key = normalized_root_path(left)
    right_key = normalized_root_path(right)
    if not left_key or not right_key:
        return False
    try:
        common = os.path.commonpath([left_key, right_key])
    except (OSError, TypeError, ValueError):
        return False
    return common == left_key or common == right_key


def root_contains(parent: object, child: object, *, include_same: bool = True) -> bool:
    """Return True when *child* is inside *parent* using lexical Windows paths."""
    parent_key = normalized_root_path(parent)
    child_key = normalized_root_path(child)
    if not parent_key or not child_key:
        return False
    try:
        common = os.path.commonpath([parent_key, child_key])
    except (OSError, TypeError, ValueError):
        return False
    if common != parent_key:
        return False
    return include_same or child_key != parent_key


@dataclass
class TaskResult:
    ok: bool
    message: str
    outputs: Dict[str, Any]


class TaskRunner:
    """UI runner wiring to verified scan/repair logic."""

    def _xlsx_to_m3u(
        self,
        xlsx_path: Path,
        out_m3u: Path,
        cancel_flag: Optional[Callable[[], bool]] = None,
    ) -> int:
        """Convert a Roon XLSX export to an internal annotated M3U.

        Each path is preceded by a #ROONMETA JSON comment so the repair engine can
        use Roon metadata even when the absolute path came from another computer.
        """
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            ws = wb.active

            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                raise ValueError("Empty xlsx")

            header_map = {str(h).strip().lower(): i for i, h in enumerate(header) if h is not None}
            if "path" not in header_map:
                raise ValueError("XLSX missing 'Path' column")

            def cell(row, *names):
                for name in names:
                    idx = header_map.get(name.lower())
                    if idx is not None and idx < len(row):
                        value = row[idx]
                        if value is not None and str(value).strip():
                            return str(value).strip()
                return ""

            lines = ["#EXTM3U", "#PLAYLISTFIXER_SOURCE:ROON_XLSX"]
            n = 0
            for r in rows:
                if cancel_flag and cancel_flag():
                    raise RepairCancelled("Repair cancelled.")
                if not r:
                    continue
                path_value = cell(r, "Path")
                if not path_value:
                    continue
                meta = {
                    "source": "roon_xlsx",
                    "album_artist": cell(r, "Album Artist"),
                    "album": cell(r, "Album"),
                    "disc": cell(r, "Disc#"),
                    "track": cell(r, "Track#"),
                    "title": cell(r, "Title"),
                    "artist": cell(r, "Track Artist(s)", "Album Artist"),
                    "external_id": cell(r, "External Id"),
                }
                lines.append("#ROONMETA:" + json.dumps(meta, ensure_ascii=False, separators=(",", ":")))
                lines.append(path_value)
                n += 1
        finally:
            wb.close()

        atomic_write_text(out_m3u, "\n".join(lines) + "\n", encoding="utf-8")
        return n

    def _coerce_playlist_to_m3u(
        self,
        pl: Path,
        out_dir: Path,
        cancel_flag: Optional[Callable[[], bool]] = None,
    ) -> tuple[Path, str]:
        """
        Convert supported playlist inputs into a m3u path we can feed to repair engine.

        - .m3u/.m3u8 : pass-through
        - .xlsx      : convert to temp m3u with path lines (preserve order)
        """
        ext = pl.suffix.lower()
        key = self.canonical_key(pl)

        if ext in (".m3u", ".m3u8"):
            return pl, pl.name

        if ext == ".xlsx":
            tmp_m3u = out_dir / f"__tmp_from_xlsx_{key}.m3u"
            self._xlsx_to_m3u(pl, tmp_m3u, cancel_flag=cancel_flag)
            return tmp_m3u, pl.name

        raise ValueError(f"Unsupported playlist type: {pl.name}")

    # -------------------------
    # Canonical key helpers
    # -------------------------
    def canonical_key(self, playlist_path: Path) -> str:
        """Return a readable key that is unique to this exact playlist file.

        Every saved output owns an independent report/selection snapshot.  The
        digest is internal metadata only and is never exposed in exported playlist
        filenames.
        """
        path = Path(playlist_path)
        stem = path.stem.strip() or "playlist"
        safe_stem = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", stem).strip(" ._") or "playlist"
        try:
            identity = str(path.resolve(strict=False)).casefold()
        except (OSError, RuntimeError):
            identity = str(path.absolute()).casefold()
        digest = hashlib.sha256(identity.encode("utf-8")).hexdigest()[:8]
        return f"{safe_stem}_{digest}"

    def report_path_for(self, out_dir: Path, playlist_path: Path) -> Path:
        key = self.canonical_key(playlist_path)
        return out_dir / f"repair_report_{key}.csv"

    def session_report_path_for(self, out_dir: Path, playlist_path: Path) -> Path:
        key = self.canonical_key(playlist_path)
        return out_dir / f"session_repair_report_{key}.csv"

    def selections_path_for(self, out_dir: Path, playlist_path: Path) -> Path:
        key = self.canonical_key(playlist_path)
        return out_dir / f"selections_{key}.json"

    def export_path_for(self, out_dir: Path, playlist_path: Path, extension: str = ".m3u8") -> Path:
        stem = Path(playlist_path).stem.strip() or "playlist"
        # Re-saving either the old 2.2.0 generated name or the new simple name
        # must not keep stacking "fixed_" prefixes.
        old_generated = re.fullmatch(r"fixed_(.+)_[0-9a-fA-F]{8}_selected", stem)
        if old_generated:
            stem = old_generated.group(1)
        elif stem.startswith("fixed_"):
            stem = stem[len("fixed_"):]
        if stem.endswith("_selected"):
            stem = stem[:-len("_selected")]
        safe_stem = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", stem).strip(" ._") or "playlist"
        ext = str(extension or ".m3u8").strip().lower()
        if ext not in (".m3u", ".m3u8"):
            ext = ".m3u8"
        return out_dir / f"fixed_{safe_stem}{ext}"

    def remove_roots_from_index(
        self,
        roots_to_remove: List[Path],
        out_index: Path,
        progress: Optional[ProgressCb] = None,
        cancel_flag: Optional[Callable[[], bool]] = None,
        preserve_roots: Optional[List[Path]] = None,
    ) -> TaskResult:
        """Remove a Music Root while handing retained child scopes their records.

        Music Roots are not only UI settings: each imported root owns records in the
        persistent music index. When a parent is removed but one or more child paths
        remain in the UI, their already indexed records are reassigned to the
        shallowest retained child instead of being discarded and rescanned.
        """
        started = time.perf_counter()
        roots = [Path(r) for r in roots_to_remove]
        requested_preserve_roots = [Path(r) for r in (preserve_roots or [])]
        norm_cache: dict[str, str] = {}

        def norm(value: object) -> str:
            text = str(value or "")
            cached = norm_cache.get(text)
            if cached is not None:
                return cached
            # Index paths are already absolute. Lexical normalization avoids the
            # filesystem access performed by Path.resolve() for every track.
            normalized = os.path.normcase(
                os.path.abspath(os.path.normpath(text))
            ).rstrip("\\/").casefold()
            norm_cache[text] = normalized
            return normalized

        removed_keys = {norm(r) for r in roots if str(r)}

        def is_within(path_key: str, root_key: str) -> bool:
            return bool(
                path_key
                and root_key
                and (
                    path_key == root_key
                    or path_key.startswith(root_key + "\\")
                    or path_key.startswith(root_key + "/")
                )
            )

        # Only descendants of a removed root can receive its records. Keep every
        # requested child here; sorting shallowest-first ensures a retained parent
        # owns its deeper retained scopes without creating overlapping index roots.
        preserve_entries: list[tuple[str, str]] = []
        seen_preserve_keys: set[str] = set()
        for preserve_root in requested_preserve_roots:
            preserve_key = norm(preserve_root)
            if (
                not preserve_key
                or preserve_key in removed_keys
                or preserve_key in seen_preserve_keys
                or not any(is_within(preserve_key, key) for key in removed_keys)
            ):
                continue
            seen_preserve_keys.add(preserve_key)
            preserve_entries.append((preserve_key, str(preserve_root)))
        preserve_entries.sort(key=lambda pair: len(pair[0]))

        if progress:
            progress(5, "Reading music index…")
        existing_items: list[dict] = []
        if out_index.exists():
            try:
                raw = json.loads(out_index.read_text(encoding="utf-8"))
                if isinstance(raw, list):
                    existing_items = [x for x in raw if isinstance(x, dict)]
            except Exception as exc:
                return TaskResult(False, f"Could not read music index: {exc}", {})

        kept_items: list[dict] = []
        removed_count = 0
        reassigned_count = 0
        total_items = max(1, len(existing_items))
        for index, item in enumerate(existing_items):
            if cancel_flag and cancel_flag():
                return TaskResult(False, "Remove cancelled.", {"cancelled": True})

            item_path_key = norm(item.get("path", ""))
            item_root_key = norm(item.get("root", ""))
            belongs_to_removed_root = any(
                is_within(item_path_key, key)
                or is_within(item_root_key, key)
                for key in removed_keys
            )
            if not belongs_to_removed_root:
                kept_items.append(item)
            else:
                new_owner = next(
                    (
                        preserve_path
                        for preserve_key, preserve_path in preserve_entries
                        if is_within(item_path_key, preserve_key)
                    ),
                    "",
                )
                if new_owner:
                    kept_items.append(dict(item, root=new_owner))
                    reassigned_count += 1
                else:
                    removed_count += 1
            if progress and index and index % 5000 == 0:
                progress(
                    10 + int(index * 55 / total_items),
                    f"Updating indexed tracks… {index:,}/{len(existing_items):,}",
                )

        ext_counts: dict[str, int] = {}
        final_roots: list[str] = []
        seen_roots: set[str] = set()
        for kept_index, item in enumerate(kept_items):
            if cancel_flag and kept_index % 1000 == 0 and cancel_flag():
                return TaskResult(False, "Remove cancelled.", {"cancelled": True})
            ext = str(item.get("ext") or Path(str(item.get("path", ""))).suffix).lower()
            if ext:
                ext_counts[ext] = ext_counts.get(ext, 0) + 1
            item_root = item.get("root")
            if item_root:
                key = norm(item_root)
                if key and key not in seen_roots:
                    seen_roots.add(key)
                    final_roots.append(str(item_root))

        stats = {
            "roots": final_roots,
            "scanned_supported": len(kept_items),
            "skipped_no_duration": 0,
            "indexed": len(kept_items),
            "available_exts": sorted(ext_counts.keys()),
            "ext_counts": ext_counts,
            "last_removed_roots": [str(r) for r in roots],
            "last_removed_indexed": removed_count,
            "last_reassigned_roots": [path for _key, path in preserve_entries],
            "last_reassigned_indexed": reassigned_count,
            "remove_seconds": round(time.perf_counter() - started, 3),
        }

        if progress:
            progress(75, "Writing updated music index…")
        atomic_write_text(
            out_index,
            json.dumps(kept_items, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )
        sp = stats_path_fn()
        atomic_write_text(sp, json.dumps(stats, ensure_ascii=False, indent=2), encoding="utf-8")
        elapsed = time.perf_counter() - started
        if progress:
            progress(100, f"Remove complete in {elapsed:.2f}s")
        return TaskResult(
            True,
            f"Removed {removed_count} and retained {reassigned_count} indexed track(s); "
            f"total library: {len(kept_items)}",
            {
                "removed_indexed": removed_count,
                "reassigned_indexed": reassigned_count,
                "reassigned_roots": [path for _key, path in preserve_entries],
                "total_indexed": len(kept_items),
                "stats": str(sp),
                "remove_seconds": round(elapsed, 3),
            },
        )

    # -------------------------
    # Scan
    # -------------------------
    def scan_index(
        self,
        music_roots: List[Path],
        out_index: Path,
        root_owners: Optional[Dict[str, str]] = None,
        progress: Optional[ProgressCb] = None,
        cancel_flag: Optional[Callable[[], bool]] = None,
    ) -> TaskResult:
        """Incrementally scan the selected roots.

        Selected roots are replaced with fresh scan results, while tracks belonging
        to unselected roots remain in the index. Removing a root is an explicit UI
        operation; leaving it unchecked during a scan must never delete it.
        """
        roots = [Path(r) for r in music_roots]
        for left_index, left_root in enumerate(roots):
            for right_root in roots[left_index + 1:]:
                if roots_overlap(left_root, right_root):
                    return TaskResult(
                        False,
                        "Cannot scan duplicate or overlapping Music Roots:\n\n"
                        f"{left_root}\n{right_root}\n\n"
                        "Remove one of these paths before scanning. "
                        "The existing music index was not changed.",
                        {
                            "overlapping_roots": [
                                str(left_root),
                                str(right_root),
                            ]
                        },
                    )
        if progress:
            progress(0, "Scanning…")

        def root_key(value: object) -> str:
            return normalized_root_path(value).rstrip("\\/")

        selected_keys = {root_key(r) for r in roots}
        owner_lookup = {
            root_key(scan_path): Path(owner_path)
            for scan_path, owner_path in (root_owners or {}).items()
            if root_key(scan_path) and str(owner_path)
        }
        owner_by_scan_key = {
            root_key(root): owner_lookup.get(root_key(root), root)
            for root in roots
        }

        # Keep entries from roots that are not part of this scan. The selected
        # roots will be replaced by the newly scanned results below.
        existing_items: list[dict] = []
        if out_index.exists():
            try:
                raw = json.loads(out_index.read_text(encoding="utf-8"))
            except Exception as exc:
                return TaskResult(
                    False,
                    "Could not read the existing music index, so Scan was stopped "
                    "without changing it.\n\n"
                    f"Index: {out_index}\n\n"
                    f"Error: {exc}\n\n"
                    "Close Playlist Fixer, move or delete the damaged index file, "
                    "then run Scan New Folders again.",
                    {"corrupt_index": str(out_index)},
                )
            if not isinstance(raw, list) or any(
                not isinstance(item, dict) for item in raw
            ):
                return TaskResult(
                    False,
                    "The existing music index has an invalid structure, so Scan "
                    "was stopped without changing it.\n\n"
                    f"Index: {out_index}\n\n"
                    "Close Playlist Fixer, move or delete the damaged index file, "
                    "then run Scan New Folders again.",
                    {"corrupt_index": str(out_index)},
                )
            existing_items = raw

        preserved_items: list[dict] = []
        selected_existing_by_root: dict[str, list[dict]] = {key: [] for key in selected_keys}
        for existing_index, item in enumerate(existing_items):
            if cancel_flag and existing_index % 1000 == 0 and cancel_flag():
                return TaskResult(False, "Scan cancelled.", {"cancelled": True})
            matched_root_key: Optional[str] = None
            item_path_key = root_key(item.get("path", ""))
            # A newly added parent root may replace several previously indexed
            # child roots. Match by physical track path first so those child
            # records are reused and replaced instead of being preserved as
            # duplicate rows with a different root owner.
            for key in selected_keys:
                if (
                    item_path_key == key
                    or item_path_key.startswith(key + "\\")
                    or item_path_key.startswith(key + "/")
                ):
                    matched_root_key = key
                    break
            if matched_root_key is None:
                item_root = item.get("root")
                if item_root:
                    candidate_root_key = root_key(item_root)
                    if candidate_root_key in selected_keys:
                        matched_root_key = candidate_root_key
            if matched_root_key is None:
                preserved_items.append(item)
            else:
                selected_existing_by_root.setdefault(matched_root_key, []).append(item)

        scanned_items: list[dict] = []
        root_results: list[dict] = []
        total = max(1, len(roots))
        scanned_supported = 0
        skipped_no_duration = 0
        reused_unchanged = 0
        parsed_new_or_changed = 0
        removed_missing = 0
        enumeration_seconds = 0.0
        parsing_seconds = 0.0
        scan_seconds = 0.0
        fast_extensions: set[str] = set()
        fallback_extensions: set[str] = set()

        for idx, r in enumerate(roots):
            if cancel_flag and cancel_flag():
                return TaskResult(False, "Scan cancelled.", {"cancelled": True})
            if progress:
                progress(
                    int(idx * 100 / total),
                    f"Scanning: {r} | indexed this scan: {len(scanned_items)} | total retained: {len(preserved_items)}",
                )

            try:
                res = scan_folder(
                    r,
                    selected_existing_by_root.get(root_key(r), []),
                    cancel_flag=cancel_flag,
                )
            except ScanCancelled:
                return TaskResult(False, "Scan cancelled.", {"cancelled": True})
            result_root = str(res.get("root", r))
            owner_root = owner_by_scan_key.get(root_key(r), r)
            result_items = [
                dict(item, root=str(owner_root))
                for item in (res.get("items", []) or [])
                if isinstance(item, dict)
            ]
            root_results.append({
                "root": str(owner_root),
                "scan_path": result_root,
                "scanned_supported": int(res.get("scanned_supported", 0) or 0),
                "skipped_no_duration": int(res.get("skipped_no_duration", 0) or 0),
                "indexed": int(res.get("indexed", 0) or 0),
                "reused_unchanged": int(res.get("reused_unchanged", 0) or 0),
                "parsed_new_or_changed": int(res.get("parsed_new_or_changed", 0) or 0),
                "removed_missing": int(res.get("removed_missing", 0) or 0),
                "fast_extensions": list(res.get("fast_extensions", []) or []),
                "fallback_extensions": list(res.get("fallback_extensions", []) or []),
                "enumeration_seconds": float(res.get("enumeration_seconds", 0) or 0),
                "parsing_seconds": float(res.get("parsing_seconds", 0) or 0),
                "total_seconds": float(res.get("total_seconds", 0) or 0),
            })
            scanned_items.extend(result_items)
            scanned_supported += int(res.get("scanned_supported", 0) or 0)
            skipped_no_duration += int(res.get("skipped_no_duration", 0) or 0)
            reused_unchanged += int(res.get("reused_unchanged", 0) or 0)
            parsed_new_or_changed += int(res.get("parsed_new_or_changed", 0) or 0)
            removed_missing += int(res.get("removed_missing", 0) or 0)
            enumeration_seconds += float(res.get("enumeration_seconds", 0) or 0)
            parsing_seconds += float(res.get("parsing_seconds", 0) or 0)
            scan_seconds += float(res.get("total_seconds", 0) or 0)
            fast_extensions.update(res.get("fast_extensions", []) or [])
            fallback_extensions.update(res.get("fallback_extensions", []) or [])

        items = preserved_items + scanned_items

        # Older versions could index one physical file through both a parent
        # root and one of its children. Keep one row per normalized path. Fresh
        # scan rows come last and replace stale preserved ownership records.
        deduplicated_paths = 0
        unique_items: list[dict] = []
        item_position_by_path: dict[str, int] = {}
        for item in items:
            path_key = root_key(item.get("path", ""))
            if not path_key:
                unique_items.append(item)
                continue
            previous_position = item_position_by_path.get(path_key)
            if previous_position is None:
                item_position_by_path[path_key] = len(unique_items)
                unique_items.append(item)
            else:
                unique_items[previous_position] = item
                deduplicated_paths += 1
        items = unique_items

        # Rebuild aggregate metadata from the final merged index so the stats file
        # describes the full library, not merely the roots scanned this time.
        ext_counts: dict[str, int] = {}
        final_roots: list[str] = []
        seen_roots: set[str] = set()
        for item_index, item in enumerate(items):
            if cancel_flag and item_index % 1000 == 0 and cancel_flag():
                return TaskResult(False, "Scan cancelled.", {"cancelled": True})
            ext = str(item.get("ext") or Path(str(item.get("path", ""))).suffix).lower()
            if ext:
                ext_counts[ext] = ext_counts.get(ext, 0) + 1
            item_root = item.get("root")
            if item_root:
                key = root_key(item_root)
                if key not in seen_roots:
                    seen_roots.add(key)
                    final_roots.append(str(item_root))

        stats = {
            "roots": final_roots,
            "scanned_supported": scanned_supported,
            "skipped_no_duration": skipped_no_duration,
            "indexed": len(items),
            "available_exts": sorted(ext_counts.keys()),
            "ext_counts": ext_counts,
            "last_scan_roots": [str(r) for r in roots],
            "last_scan_owners": {
                str(root): str(owner_by_scan_key.get(root_key(root), root))
                for root in roots
            },
            "last_scan_indexed": len(scanned_items),
            "preserved_indexed": len(preserved_items),
            "reused_unchanged": reused_unchanged,
            "parsed_new_or_changed": parsed_new_or_changed,
            "removed_missing": removed_missing,
            "deduplicated_paths": deduplicated_paths,
            "fast_extensions": sorted(fast_extensions),
            "fallback_extensions": sorted(fallback_extensions),
            "scan_timing_seconds": {
                "enumeration": round(enumeration_seconds, 3),
                "parsing": round(parsing_seconds, 3),
                "total": round(scan_seconds, 3),
            },
        }

        atomic_write_text(
            out_index,
            json.dumps(items, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )

        sp = stats_path_fn()
        atomic_write_text(sp, json.dumps(stats, ensure_ascii=False, indent=2), encoding="utf-8")

        if progress:
            progress(
                100,
                f"Scan complete. Parsed: {parsed_new_or_changed} | Reused: {reused_unchanged} | "
                f"Removed: {removed_missing} | Total indexed: {len(items)}",
            )
        return TaskResult(
            True,
            f"Scan complete. Added/updated: {len(scanned_items)} | Total indexed: {len(items)}",
            {
                "index": str(out_index),
                "stats": str(sp),
                "root_results": root_results,
                "scan_indexed": len(scanned_items),
                "total_indexed": len(items),
                "preserved_indexed": len(preserved_items),
                "deduplicated_paths": deduplicated_paths,
            },
        )

    # -------------------------
    # Report helpers
    # -------------------------
    def _read_report_rows(self, report_csv: Path) -> list[dict]:
        rows: list[dict] = []
        if not report_csv.exists():
            return rows
        with report_csv.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for i, r in enumerate(reader):
                r["_i"] = i
                if "row_index" not in r or r["row_index"] in (None, ""):
                    r["row_index"] = str(i)
                else:
                    r["row_index"] = str(r["row_index"]).strip()
                rows.append(r)
        return rows

    def _parse_candidates_from_notes(self, notes: str) -> list[str]:
        """
        Parse candidate file paths from notes.

        - Prefer part after 'candidates:' if present.
        - Split by '|'
        - Keep only tokens that look like a file path (has separator + filename contains a dot)
        """
        notes = (notes or "").strip()
        if not notes:
            return []

        low = notes.lower()
        if "candidates:" in low:
            notes = notes[low.index("candidates:") + len("candidates:") :].strip()

        cands: list[str] = []
        for part in notes.split("|"):
            p = (part or "").strip().strip('"').strip("'")
            if not p:
                continue

            has_sep = (":\\" in p) or ("/" in p) or ("\\" in p)
            name = Path(p).name
            looks_like_file = ("." in name) and len(name) >= 3

            if has_sep and looks_like_file:
                cands.append(p)

        return cands

    def _picked_path_from_row(self, r: dict) -> str:
        """
        Detect "final chosen path" from report columns.
        Do NOT use notes.
        """
        if not isinstance(r, dict):
            return ""

        KEY_HINTS = (
            "written", "written_path",
            "chosen", "chosen_path",
            "selected", "selected_path",
            "best", "best_path", "best_match", "best_match_path",
            "final", "final_path",
            "resolved", "resolved_path",
            "output", "output_path",
            "matched", "matched_path",
            "picked", "picked_path",
            "target", "target_path",
            "result", "result_path",
        )

        for k, v in r.items():
            lk = str(k).lower()
            if any(h in lk for h in KEY_HINTS):
                s = (str(v) if v is not None else "").strip()
                if s:
                    return s

        return ""

    def _classify_for_ui(self, report_rows: list[dict], playlist_path: Path) -> tuple[list[dict], list[dict]]:
        """
        UNRESOLVED UI classification (robust):
        - Show only rows that still need human action.
        - Skip resolved rows.
        - Robust to different status naming conventions by using keyword matching.
        - Do NOT fallback unknown statuses into FAILED blindly; only classify if it matches patterns.
        """
        ambiguous: list[dict] = []
        failed: list[dict] = []

        pl_key = self.canonical_key(playlist_path)

        def norm(s: str) -> str:
            return (s or "").strip().upper()

        def classify_status(st: str) -> str:
            """
            Return: 'RESOLVED' | 'AMBIGUOUS' | 'FAILED' | ''(unknown)
            """
            st = norm(st)
            if not st:
                return ""

            # resolved keywords
            if any(k in st for k in ("KEPT", "REPAIRED", "FIXED", "OK", "DONE", "SUCCESS", "RESOLV")):
                return "RESOLVED"

            # ambiguous-ish keywords
            if any(k in st for k in ("AMBIG", "MULTI", "CONFLICT", "DUPLIC", "CANDIDATE", "MULTIPLE")):
                return "AMBIGUOUS"

            # failed-ish keywords
            if any(k in st for k in ("FAIL", "NOT_FOUND", "NOTFOUND", "MISSING", "MISS", "ERROR", "ERR")):
                return "FAILED"

            return ""

        for r in report_rows:
            status_raw = r.get("status") or ""
            kind = classify_status(status_raw)

            # resolved => never show in unresolved lists
            if kind == "RESOLVED":
                continue

            # only show known unresolved kinds
            if kind not in ("AMBIGUOUS", "FAILED"):
                continue

            row_index_raw = (r.get("row_index") or r.get("_i") or -1)
            try:
                row_index = int(str(row_index_raw).strip())
            except Exception:
                row_index = -1

            extinf_display = (r.get("extinf_display") or r.get("extinf") or "").strip()
            notes = (r.get("notes") or "").strip()
            orig = (r.get("original_path") or r.get("original") or "").strip()

            cands = self._parse_candidates_from_notes(notes)

            row = {
                "playlist": str(playlist_path),
                "pl_key": pl_key,
                "row_index": row_index,
                "extinf_display": extinf_display,
                "original_path": orig,
                "notes": notes,
                "candidates": cands,
            }

            if kind == "AMBIGUOUS":
                ambiguous.append(row)
            else:
                failed.append(row)

        return ambiguous, failed

    # -------------------------
    # Repair phase (NO playlist output kept)
    # -------------------------
    def repair_playlists(
        self,
        playlists: List[Path],
        index_path: Path,
        out_dir: Path,
        mode: str = "safe",
        dry_run: bool = False,
        progress: Optional[ProgressCb] = None,
        cancel_flag: Optional[Callable[[], bool]] = None,
        format_mode: str = "none",
        strict_ext: Optional[str] = None,
        format_priority_list: Optional[List[str]] = None,
        enabled_roots: Optional[List[Path]] = None,
        session_reports: bool = False,
    ) -> TaskResult:
        """
        Repair(Safe):
        - Generates repair_report_{key}.csv
        - Does NOT keep auto-generated fixed playlist file (tmp file deleted).
        """
        out_dir.mkdir(parents=True, exist_ok=True)

        # Pass the complete persistent index to the repair engine.
        #
        # The engine itself restricts repair candidates to ``enabled_roots`` via
        # ``allowed_roots``. Keeping the complete index here is intentional: it
        # lets the engine look up metadata for the exact original playlist path
        # (title/artist/album/track/duration) while ensuring files outside the
        # selected roots can never become repair candidates.
        repair_index_path = index_path

        summaries: list[dict] = []
        all_amb: list[dict] = []
        all_fail: list[dict] = []

        total_pl = max(1, len(playlists))
        for i, pl in enumerate(playlists):
            if cancel_flag and cancel_flag():
                return TaskResult(False, "Repair cancelled.", {"cancelled": True})

            pct = int(i * 100 / total_pl)
            if progress:
                progress(pct, f"Repairing: {pl.name}")

            key = self.canonical_key(pl)
            m3u_in: Optional[Path] = None
            tmp_fixed = out_dir / f"__tmp_fixed_{key}.m3u"
            report_path = (self.session_report_path_for(out_dir, pl)
                           if session_reports else self.report_path_for(out_dir, pl))
            try:
                m3u_in, _ = self._coerce_playlist_to_m3u(
                    pl,
                    out_dir,
                    cancel_flag=cancel_flag,
                )
                s = repair_playlist(
                    str(m3u_in),
                    str(repair_index_path),
                    str(tmp_fixed),
                    str(report_path),
                    verbose=False,
                    format_mode=format_mode,
                    strict_ext=strict_ext,
                    format_priority_list=format_priority_list,
                    allowed_roots=[str(r) for r in enabled_roots] if enabled_roots else None,
                    cancel_flag=cancel_flag,
                    source_base_dir=str(pl.parent),
                )
                summaries.append(s)
            except RepairCancelled:
                return TaskResult(False, "Repair cancelled.", {"cancelled": True})
            finally:
                # Conversion and repair output are internal only. Always remove
                # them after success, failure, or cancellation.
                try:
                    if (
                        m3u_in is not None
                        and m3u_in.name.startswith("__tmp_from_xlsx_")
                        and m3u_in.exists()
                    ):
                        m3u_in.unlink()
                except OSError:
                    pass
                try:
                    if tmp_fixed.exists():
                        tmp_fixed.unlink()
                except OSError:
                    pass

            report_rows = self._read_report_rows(report_path)
            if report_rows:
                amb, fail = self._classify_for_ui(report_rows, pl)
                all_amb.extend(amb)
                all_fail.extend(fail)

        if progress:
            progress(100, "Repair complete.")
        return TaskResult(
            True,
            "Repair complete.",
            {
                "summaries": summaries,
                "out_dir": str(out_dir),
                "ambiguous": all_amb,
                "failed": all_fail,
            },
        )

    def _xlsx_export_metadata(
        self,
        xlsx_path: Path,
        cancel_flag: Optional[Callable[[], bool]] = None,
    ) -> Dict[str, dict]:
        """Return row-indexed reliable metadata from a Roon XLSX export.

        The row indexes match the playlist/report order (0-based, excluding the
        header).  This is used only when exporting an XLSX-derived repair result
        to M3U, so the structured Title/Artist data is not lost.
        """
        result: Dict[str, dict] = {}
        if xlsx_path.suffix.lower() != ".xlsx" or not xlsx_path.exists():
            return result

        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                return result
            header_map = {str(h).strip().lower(): i for i, h in enumerate(header) if h is not None}

            def cell(row, *names):
                for name in names:
                    idx = header_map.get(name.lower())
                    if idx is not None and idx < len(row):
                        value = row[idx]
                        if value is not None and str(value).strip():
                            return str(value).strip()
                return ""

            output_index = 0
            for row in rows:
                if cancel_flag and cancel_flag():
                    raise RepairCancelled("Export cancelled.")
                if not row:
                    continue
                path_value = cell(row, "Path")
                if not path_value:
                    continue

                artist = cell(row, "Track Artist(s)", "Album Artist")
                # Roon sometimes exports the same artist twice, e.g.
                # "NEFFEX / NEFFEX". Remove exact duplicates without changing
                # genuine collaborations.
                if artist:
                    parts = [part.strip() for part in artist.split(" / ") if part.strip()]
                    unique_parts = []
                    seen = set()
                    for part in parts:
                        key = part.casefold()
                        if key not in seen:
                            seen.add(key)
                            unique_parts.append(part)
                    artist = " / ".join(unique_parts)

                result[str(output_index)] = {
                    "title": cell(row, "Title"),
                    "artist": artist,
                    "album_artist": cell(row, "Album Artist"),
                    "album": cell(row, "Album"),
                    "disc": cell(row, "Disc#"),
                    "track": cell(row, "Track#"),
                    "original_path": path_value,
                }
                output_index += 1
        finally:
            wb.close()
        return result

    # -------------------------
    # Save/Export phase (final playlist output)
    # -------------------------
    def export_fixed_multi(
        self,
        jobs: List[Dict[str, Any]],
        progress: Optional[ProgressCb] = None,
        cancel_flag: Optional[Callable[[], bool]] = None,
    ) -> TaskResult:
        """
        Save/export final playlists.

        - Use STRICT final-path whitelist keys to avoid accidentally using candidates columns.
        - Manual selections override everything.
        - For resolved statuses, write final path (from whitelist) or orig if missing.
        - For unresolved statuses, keep original path.
        """

        FINAL_KEYS = (
            "written_path", "written",
            "final_path", "final",
            "resolved_path", "resolved",
            "picked_path", "picked",
            "chosen_path", "chosen",
            "selected_path", "selected",
            "output_path", "output",
            "result_path", "result",
            "target_path", "target",
            "matched_path", "matched",
        )

        RESOLVED_KEYWORDS = ("KEPT", "REPAIRED", "FIXED", "OK", "DONE", "SUCCESS", "RESOLV")

        def is_resolved_status(st: str) -> bool:
            st = (st or "").strip().upper()
            return any(k in st for k in RESOLVED_KEYWORDS)

        def pick_final(rr: dict) -> str:
            for k in FINAL_KEYS:
                v = rr.get(k)
                s = (str(v) if v is not None else "").strip()
                if s:
                    return s
            return ""

        done = []
        prepared_outputs: list[tuple[Path, str, str]] = []
        total = max(1, len(jobs))

        for i, job in enumerate(jobs):
            if cancel_flag and cancel_flag():
                return TaskResult(False, "Export cancelled.", {"cancelled": True})

            if progress:
                progress(int(i * 100 / total), f"Saving: {Path(job['out_m3u']).name}")

            report_csv = Path(job["report_csv"])
            out_m3u = Path(job["out_m3u"])
            selections: Dict[str, str] = {
                str(k).strip(): str(v)
                for k, v in (job.get("selections", {}) or {}).items()
            }
            selection_records = job.get("selection_records", []) or []

            def normalized_row_id(value: object) -> str:
                text = str(value if value is not None else "").strip()
                if not text:
                    return ""
                try:
                    number = float(text)
                    if number.is_integer():
                        return str(int(number))
                except (TypeError, ValueError):
                    pass
                return text

            # Normalize ids and keep an original-path fallback.  Manual choices
            # must always override report output, including Failed and Resolved
            # rows and Ambiguous choices other than candidate 1.
            selections_by_id = {normalized_row_id(k): v for k, v in selections.items()}
            selections_by_original: Dict[str, str] = {}
            for record in selection_records:
                if not isinstance(record, dict):
                    continue
                chosen_path = str(record.get("chosen_path", "")).strip()
                original_path = str(record.get("original_path", "")).strip()
                row_id = normalized_row_id(record.get("row_index", ""))
                if chosen_path and row_id:
                    selections_by_id[row_id] = chosen_path
                if chosen_path and original_path:
                    selections_by_original[original_path.casefold()] = chosen_path

            # Build a path -> indexed metadata lookup once per export job.
            # This lets us enrich missing or incomplete standard EXTINF data without
            # reopening every audio file during Save. Existing good EXTINF text is
            # preserved exactly.
            metadata_by_path: Dict[str, dict] = {}
            index_path_value = str(job.get("index_path", "") or "").strip()
            if index_path_value:
                try:
                    raw_index = json.loads(Path(index_path_value).read_text(encoding="utf-8"))
                    if isinstance(raw_index, list):
                        for item in raw_index:
                            if not isinstance(item, dict):
                                continue
                            path_value = str(item.get("path", "") or "").strip()
                            if path_value:
                                metadata_by_path[path_value.casefold()] = item
                except Exception:
                    metadata_by_path = {}

            # Roon XLSX has structured metadata that ordinary M3U files do not.
            # Preserve it when exporting to M3U instead of reducing the result to
            # path-only lines.  Ordinary M3U/M3U8 behavior remains unchanged.
            xlsx_metadata_by_row: Dict[str, dict] = {}
            source_playlist_value = str(job.get("source_playlist", "") or "").strip()
            if source_playlist_value:
                try:
                    source_playlist = Path(source_playlist_value)
                    if source_playlist.suffix.lower() == ".xlsx":
                        xlsx_metadata_by_row = self._xlsx_export_metadata(
                            source_playlist,
                            cancel_flag=cancel_flag,
                        )
                except RepairCancelled:
                    return TaskResult(False, "Export cancelled.", {"cancelled": True})
                except Exception:
                    xlsx_metadata_by_row = {}

            extinf_re = re.compile(r"^#EXTINF:\s*(-?\d+)\s*,\s*(.*)$", re.IGNORECASE)

            def build_extinf(existing: str, output_path: str, row_index: str = "") -> str:
                """Preserve good original EXTINF; fill only missing facts.

                Rules:
                - A valid EXTINF with a non-negative duration and non-empty display
                  is returned unchanged.
                - If only duration is missing/negative, keep the original display
                  and fill duration from indexed metadata when available.
                - If EXTINF is absent or display text is empty, construct standard
                  Artist - Title text only from reliable indexed metadata.
                - Never invent artist/title. If metadata is insufficient, fall back
                  to the file stem as display text.
                """
                existing = (existing or "").strip()
                metadata = metadata_by_path.get((output_path or "").casefold(), {})
                xlsx_metadata = xlsx_metadata_by_row.get(normalized_row_id(row_index), {})
                duration = metadata.get("duration")
                try:
                    duration_int = int(round(float(duration))) if duration not in (None, "") else None
                except (TypeError, ValueError):
                    duration_int = None

                title = str(metadata.get("title", "") or "").strip()
                artist = str(metadata.get("artist", "") or "").strip()

                # If the repaired target is outside the active index, retain the
                # trustworthy metadata supplied by the original Roon XLSX.
                if not title:
                    title = str(xlsx_metadata.get("title", "") or "").strip()
                if not artist:
                    artist = str(
                        xlsx_metadata.get("artist", "")
                        or xlsx_metadata.get("album_artist", "")
                        or ""
                    ).strip()

                match = extinf_re.match(existing) if existing else None
                if match:
                    old_duration = int(match.group(1))
                    display = match.group(2).strip()
                    if old_duration >= 0 and display:
                        return existing
                    if display:
                        effective_duration = duration_int if duration_int is not None else old_duration
                        return f"#EXTINF:{effective_duration},{display}"

                if artist and title:
                    display = f"{artist} - {title}"
                elif title:
                    display = title
                else:
                    display = Path(output_path).stem if output_path else ""

                if not display:
                    return existing
                effective_duration = duration_int if duration_int is not None else -1
                return f"#EXTINF:{effective_duration},{display}"

            rows = self._read_report_rows(report_csv)
            lines = ["#EXTM3U"]

            for r in rows:
                if cancel_flag and cancel_flag():
                    return TaskResult(False, "Export cancelled.", {"cancelled": True})
                row_index = normalized_row_id(r.get("row_index", r.get("_i", "")))
                extinf_line = (r.get("extinf_line") or r.get("extinf") or "").strip()
                orig = (r.get("original_path") or r.get("original") or "").strip()
                status = (r.get("status") or "").strip()

                final_path = pick_final(r)
                chosen = selections_by_id.get(row_index) if row_index else None
                if not chosen and orig:
                    chosen = selections_by_original.get(orig.casefold())

                if chosen:
                    output_path = chosen
                elif is_resolved_status(status):
                    output_path = final_path or orig
                else:
                    output_path = orig

                enriched_extinf = build_extinf(extinf_line, output_path, row_index)
                if enriched_extinf:
                    lines.append(enriched_extinf)
                lines.append(output_path)

            # .m3u8 explicitly denotes UTF-8. For legacy .m3u, write a UTF-8
            # BOM so older players have a better chance of detecting Unicode paths.
            output_encoding = "utf-8-sig" if out_m3u.suffix.lower() == ".m3u" else "utf-8"
            prepared_outputs.append(
                (out_m3u, "\n".join(lines) + "\n", output_encoding)
            )

        # Do not leave a half-saved group when cancellation happens while
        # processing several playlists. Commit only after every output is ready.
        if cancel_flag and cancel_flag():
            return TaskResult(False, "Export cancelled.", {"cancelled": True})

        for out_m3u, output_text, output_encoding in prepared_outputs:
            if progress:
                progress(95, f"Writing: {out_m3u.name}")
            atomic_write_text(
                out_m3u,
                output_text,
                encoding=output_encoding,
            )
            done.append({"out_m3u": str(out_m3u)})

        if progress:
            progress(100, "Save complete.")
        return TaskResult(True, "Save complete.", {"done": done})
