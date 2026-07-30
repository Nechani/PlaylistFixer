from __future__ import annotations

import os
import re
import json
import shutil
from datetime import datetime, timezone
from pathlib import Path

from PySide6.QtCore import QThread, Signal, QObject, Slot, Qt
from openpyxl import load_workbook

from PySide6.QtWidgets import (
    QMainWindow, QWidget, QFileDialog, QMessageBox, QCheckBox, QComboBox,
    QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QProgressBar, QTableWidget, QTableWidgetItem, QGroupBox,
    QListWidget, QListWidgetItem, QAbstractItemView,
    QLineEdit, QDialog, QTextBrowser, QSplitter, QListView, QTreeView, QInputDialog,
    QSizePolicy
)

from core.runner import (
    TaskRunner,
    TaskResult,
    atomic_write_text,
    normalized_root_path,
    root_contains,
)
from core.paths import index_path, reports_dir, settings_path, stats_path
from core.version import APP_VERSION
from ui.about_dialog import AboutDialog


class Worker(QObject):
    progress = Signal(int, str)
    finished = Signal(object)
    failed = Signal(str)

    def __init__(self, func, kwargs):
        super().__init__()
        self.func = func
        self.kwargs = kwargs
        self._cancelled = False

    def cancel(self):
        self._cancelled = True

    @Slot()
    def run(self):
        try:
            result = self.func(
                progress=self.progress.emit,
                cancel_flag=lambda: self._cancelled,
                **self.kwargs
            )
            self.finished.emit(result)
        except Exception as e:
            self.failed.emit(str(e))


class MainWindow(QMainWindow):
    def __init__(self, app_data=None):
        super().__init__()
        self.app_data = app_data
        self.setWindowTitle(f"Playlist Fixer v{APP_VERSION}")

        self.runner = TaskRunner()

        self.music_roots: list[dict] = []
        self.playlists: list[Path] = []

        self.index_path = index_path()
        self.reports_path = reports_dir()

        # rows currently displayed (depends on view mode)
        self._ambiguous_rows: list[dict] = []
        self._failed_rows: list[dict] = []
        # master rows (unfiltered)
        self._ambiguous_rows_all: list[dict] = []
        self._failed_rows_all: list[dict] = []


        # maps for quick lookup: "{pl_key}::{row_id}"
        self._amb_by_id: dict[str, dict] = {}
        self._fail_by_id: dict[str, dict] = {}

        # selections keyed by stable pl_key -> {row_index(str): chosen_path}
        self._selections_by_key: dict[str, dict[str, str]] = {}
        # Rows changed by Apply in the current unsaved session.
        # Kept separate from persisted selections so only new manual edits show ✓.
        self._dirty_selection_ids: set[tuple[str, str]] = set()
        # Remember where an unsaved Apply was made.  An edit made in
        # Unresolved must stay only in Unresolved until Save; an edit made in
        # Resolved may be previewed in Resolved immediately.
        self._selection_origin: dict[tuple[str, str], str] = {}

        # cache report rows per playlist key (raw report csv rows)
        self._report_rows_by_key: dict[str, list[dict]] = {}
        self._session_repaired_keys: set[str] = set()
        
        self._saved_keys: set[str] = set()
        self._persisted_progress_keys: set[str] = set()
        # Imported fixed playlists may be moved together with their report,
        # selection and progress sidecars.  Their canonical key changes because
        # it contains the absolute path, so remember the actual recovered files.
        self._persisted_artifacts_by_key: dict[str, dict[str, Path]] = {}
        # Playlists currently being re-repaired provisionally from a clean state.
        # Existing saved reports/selections remain on disk until Save, but must not
        # leak into the current Unresolved/Resolved views.
        self._provisional_reset_keys: set[str] = set()

        self._active_target: str | None = None   # "AMBIGUOUS" | "FAILED"
        self._active_pl_key: str | None = None
        self._active_row_id: str | None = None

        self._busy = False
        self._last_progress_msg = ""

        self.thread: QThread | None = None
        self.worker: Worker | None = None

        self._pending_save_keys: list[str] = []
        self._pending_save_snapshots: list[dict] = []
        self._pending_remove_paths: set[str] = set()
        self._pending_repair_keys: set[str] = set()
        self._last_action: str | None = None
        self._close_when_finished = False
        self._close_without_prompt = False

        # view mode: "UNRESOLVED" | "RESOLVED"
        self._view_mode: str = "UNRESOLVED"

        self._build_ui()
        self._load_music_roots()
        self._refresh_music_roots_ui()

    # ---------- persistence ----------
    def _selection_file_for_key(self, pl_key: str) -> Path:
        return self.reports_path / f"selections_{pl_key}.json"

    def _load_selections_for_key(self, pl_key: str) -> dict[str, str]:
        artifacts = self._persisted_artifacts_by_key.get(pl_key, {})
        p = artifacts.get("selections", self._selection_file_for_key(pl_key))
        if not p.exists():
            return {}
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                return {str(k): str(v) for k, v in data.items()}
        except Exception:
            pass
        return {}

    def _save_selections_for_key(self, pl_key: str, sel: dict[str, str]) -> None:
        p = self._selection_file_for_key(pl_key)
        atomic_write_text(
            p,
            json.dumps(sel, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def _progress_file_for_key(self, pl_key: str) -> Path:
        return self.reports_path / f"progress_{pl_key}.json"

    @staticmethod
    def _normalized_playlist_path(path: Path) -> str:
        try:
            return str(Path(path).resolve(strict=False)).casefold()
        except Exception:
            return str(Path(path).absolute()).casefold()

    def _saved_progress_artifacts(
        self,
        pl_key: str,
        playlist: Path,
    ) -> dict[str, Path] | None:
        """Locate a saved report bundle even after its folder was moved.

        A user may save the same repair result as both M3U8 and M3U.  Version 3
        markers retain every output path instead of replacing the previous format.
        Earlier builds keyed sidecars by the output's absolute path, so moving the
        fixed playlist and all of its sidecars changed the calculated key and made
        intact progress look missing.  Prefer an exact path match, then accept a
        same-folder sidecar whose recorded output filename matches.
        """
        playlist = Path(playlist)
        exact_marker = self._progress_file_for_key(pl_key)
        marker_candidates: list[Path] = [exact_marker]
        try:
            marker_candidates.extend(playlist.parent.glob("progress_*.json"))
        except Exception:
            pass
        try:
            marker_candidates.extend(self.reports_path.rglob("progress_*.json"))
        except Exception:
            pass

        seen_markers: set[str] = set()
        matches: list[tuple[int, float, dict[str, Path]]] = []
        wanted_path = self._normalized_playlist_path(playlist)
        wanted_parent = self._normalized_playlist_path(playlist.parent)
        wanted_name = playlist.name.casefold()

        for marker in marker_candidates:
            marker_key = self._normalized_playlist_path(marker)
            if marker_key in seen_markers or not marker.exists():
                continue
            seen_markers.add(marker_key)
            try:
                data = json.loads(marker.read_text(encoding="utf-8"))
                if not isinstance(data, dict):
                    continue
            except Exception:
                continue

            saved_paths = data.get("saved_playlists")
            if not isinstance(saved_paths, list):
                saved_paths = []
            legacy_path = data.get("saved_playlist")
            if legacy_path:
                saved_paths.append(legacy_path)
            saved_paths = [str(path) for path in saved_paths if path]
            exact_path_match = any(
                self._normalized_playlist_path(Path(saved_path)) == wanted_path
                for saved_path in saved_paths
            )
            filename_match = any(
                Path(saved_path).name.casefold() == wanted_name
                for saved_path in saved_paths
            )
            same_parent = (
                self._normalized_playlist_path(marker.parent) == wanted_parent
            )
            exact_key_marker = (
                self._normalized_playlist_path(marker)
                == self._normalized_playlist_path(exact_marker)
            )

            # A filename match on its own is used only as a unique last resort.
            if exact_path_match:
                priority = 0
            # A key match must not make an M3U marker count as progress for
            # a different M3U8 output (or vice versa).  Markers that predate
            # saved path metadata may still use the key-only fallback.
            elif exact_key_marker and (not saved_paths or filename_match):
                priority = 1
            elif same_parent and filename_match:
                priority = 2
            elif filename_match:
                priority = 3
            else:
                continue

            stored_key = str(data.get("playlist_key") or "").strip()
            if not stored_key:
                stem = marker.stem
                stored_key = stem[len("progress_"):] if stem.startswith("progress_") else ""
            if not stored_key:
                continue
            report = marker.parent / f"repair_report_{stored_key}.csv"
            if not report.exists():
                continue
            selections = marker.parent / f"selections_{stored_key}.json"
            artifacts = {
                "progress": marker,
                "report": report,
                "selections": selections,
            }
            try:
                modified = marker.stat().st_mtime
            except OSError:
                modified = 0.0
            matches.append((priority, modified, artifacts))

        if not matches:
            return None
        best_priority = min(priority for priority, _modified, _artifacts in matches)
        best = [entry for entry in matches if entry[0] == best_priority]
        # A basename-only match is unsafe when several independent saved outputs
        # have the same filename. Same-folder and exact matches remain unambiguous.
        if best_priority == 3 and len(best) != 1:
            return None
        best.sort(key=lambda entry: entry[1], reverse=True)
        return best[0][2]

    def _has_saved_progress(self, pl_key: str, playlist: Path) -> bool:
        artifacts = self._saved_progress_artifacts(pl_key, playlist)
        if artifacts is None:
            self._persisted_artifacts_by_key.pop(pl_key, None)
            return False
        self._persisted_artifacts_by_key[pl_key] = artifacts
        return True

    def _write_progress_marker(self, pl_key: str, saved_playlist: Path, source_playlist: Path) -> None:
        p = self._progress_file_for_key(pl_key)
        existing: dict = {}
        if p.exists():
            try:
                loaded = json.loads(p.read_text(encoding="utf-8"))
                if isinstance(loaded, dict):
                    existing = loaded
            except Exception:
                existing = {}

        saved_paths = existing.get("saved_playlists")
        if not isinstance(saved_paths, list):
            saved_paths = []
        legacy_path = existing.get("saved_playlist")
        if legacy_path:
            saved_paths.append(legacy_path)
        saved_paths.append(str(saved_playlist))

        unique_saved_paths: list[str] = []
        seen: set[str] = set()
        for saved_path in saved_paths:
            if not saved_path:
                continue
            normalized = self._normalized_playlist_path(Path(saved_path))
            if normalized not in seen:
                seen.add(normalized)
                unique_saved_paths.append(str(saved_path))

        data = {
            "version": 3,
            "playlist_key": pl_key,
            # Keep the latest value for backward compatibility with 2.2.0.
            "saved_playlist": str(saved_playlist),
            "saved_playlists": unique_saved_paths,
            # Never replace the true original with a previously exported file.
            "source_playlist": existing.get("source_playlist") or str(source_playlist),
            "saved_at_utc": datetime.now(timezone.utc).isoformat(),
        }
        atomic_write_text(
            p,
            json.dumps(data, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def _commit_saved_snapshot(self, snapshot: dict) -> str:
        """Persist an independent report and selection snapshot for one output."""
        saved_output = Path(snapshot["saved_output"])
        output_key = self.runner.canonical_key(saved_output)
        source_report = Path(snapshot["report_csv"])
        output_report = self.runner.report_path_for(self.reports_path, saved_output)
        if (
            source_report.exists()
            and self._normalized_playlist_path(source_report)
            != self._normalized_playlist_path(output_report)
        ):
            output_report.parent.mkdir(parents=True, exist_ok=True)
            tmp_report = output_report.with_name(
                f".{output_report.name}.{os.getpid()}.tmp"
            )
            try:
                shutil.copy2(source_report, tmp_report)
                os.replace(tmp_report, output_report)
            finally:
                try:
                    if tmp_report.exists():
                        tmp_report.unlink()
                except OSError:
                    pass
        self._save_selections_for_key(output_key, snapshot.get("selections", {}) or {})
        self._write_progress_marker(
            output_key,
            saved_output,
            Path(snapshot["original_source"]),
        )
        return output_key

    def _session_report_for(self, playlist: Path) -> Path:
        return self.runner.session_report_path_for(self.reports_path, playlist)

    def _settings_file(self) -> Path:
        # Keep user preferences outside the extracted program folder so an
        # update/re-extract does not erase saved Music Roots.
        if self.app_data:
            return Path(self.app_data) / "settings.json"
        return settings_path()

    def _load_settings(self) -> dict:
        p = self._settings_file()

        # One-time migration from builds that stored settings beside app.py.
        legacy = settings_path()
        if not p.exists() and legacy.exists() and legacy != p:
            try:
                atomic_write_text(
                    p,
                    legacy.read_text(encoding="utf-8"),
                    encoding="utf-8",
                )
            except Exception:
                pass

        if not p.exists():
            return {}
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}

    def _save_settings(self, data: dict) -> None:
        p = self._settings_file()
        try:
            atomic_write_text(
                p,
                json.dumps(data, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        except Exception as e:
            QMessageBox.warning(self, "Settings not saved", f"Could not save settings to:\n{p}\n\n{e}")


    def _load_music_roots(self) -> None:
        settings = self._load_settings()
        raw = settings.get("music_roots", [])
        indexed = self._indexed_root_paths()

        def norm(value: object) -> str:
            return os.path.normcase(os.path.normpath(str(value)))

        indexed_by_key = {norm(x): str(x) for x in indexed}
        roots: list[dict] = []
        seen: set[str] = set()

        # Settings are the source of truth for what the user added.  Do not
        # delete entries merely because the index or stats file is temporarily
        # missing, unreadable, or belongs to a disconnected external drive.
        if isinstance(raw, list):
            for entry in raw:
                if isinstance(entry, str):
                    path = entry
                    enabled = True
                    saved_imported = True
                elif isinstance(entry, dict) and entry.get("path"):
                    path = str(entry["path"])
                    enabled = bool(entry.get("enabled", True))
                    saved_imported = bool(entry.get("imported", False))
                    scope_only = bool(entry.get("scope_only", False))
                    covered_by = str(entry.get("covered_by", "") or "")
                else:
                    continue
                if isinstance(entry, str):
                    scope_only = False
                    covered_by = ""

                key = norm(path)
                if not key or key in seen:
                    continue
                seen.add(key)
                imported = saved_imported or key in indexed_by_key
                roots.append({
                    "path": path,
                    "enabled": enabled,
                    # Keep the saved state, but also recover Imported when the
                    # stats file still confirms this root.
                    "imported": imported,
                    # A saved imported root without a matching stats entry means
                    # the index metadata was deleted, damaged, or replaced.
                    "index_missing": imported and key not in indexed_by_key,
                    "scope_only": scope_only,
                    "covered_by": covered_by,
                })

        # Music Roots are computer-specific settings stored in AppData.
        # Index metadata may have been carried from another computer with the
        # portable program folder, so it must never create additional roots.
        # Roots found only in music_index.stats.json are intentionally ignored.

        self.music_roots = roots
        self._reconcile_scope_roots(indexed)

    def _nearest_covering_root_entry(
        self,
        path: object,
        entries: list[dict] | None = None,
    ) -> dict | None:
        """Return the closest non-scope Music Root that contains *path*."""
        candidates = []
        for entry in entries if entries is not None else self.music_roots:
            entry_path = entry.get("path")
            if (
                entry_path
                and not entry.get("scope_only", False)
                and root_contains(entry_path, path, include_same=False)
            ):
                candidates.append(entry)
        if not candidates:
            return None
        return max(
            candidates,
            key=lambda entry: len(normalized_root_path(entry.get("path", ""))),
        )

    def _new_music_root_entry(self, path: Path) -> dict:
        """Create an index root or a scope that reuses an existing parent index."""
        covering = self._nearest_covering_root_entry(path)
        if covering is not None:
            usable = bool(
                covering.get("imported", False)
                and not covering.get("index_missing", False)
            )
            return {
                "path": str(path),
                "enabled": True,
                "imported": usable,
                "index_missing": not usable,
                "scope_only": True,
                "covered_by": str(covering.get("path", "")),
            }
        return {
            "path": str(path),
            "enabled": True,
            "imported": False,
            "index_missing": False,
            "scope_only": False,
            "covered_by": "",
        }

    def _reconcile_scope_roots(self, indexed_paths: set[str] | None = None) -> None:
        """Keep scope aliases attached to one real index owner.

        A child of an indexed root reuses that root's records. If its covering
        root disappears, the shallowest orphan becomes a Pending index root and
        deeper aliases attach to it.
        """
        indexed_values = indexed_paths if indexed_paths is not None else self._indexed_root_paths()
        indexed_keys = {
            normalized_root_path(path)
            for path in indexed_values
            if normalized_root_path(path)
        }
        ordered = sorted(
            self.music_roots,
            key=lambda entry: len(normalized_root_path(entry.get("path", ""))),
        )
        actual_entries: list[dict] = []
        scope_entries: list[dict] = []

        for entry in ordered:
            entry.setdefault("scope_only", False)
            entry.setdefault("covered_by", "")
            key = normalized_root_path(entry.get("path", ""))
            if not key:
                continue
            if entry.get("scope_only", False):
                scope_entries.append(entry)
                continue

            is_indexed = key in indexed_keys
            indexed_cover = self._nearest_covering_root_entry(
                entry.get("path", ""),
                [
                    candidate
                    for candidate in actual_entries
                    if candidate.get("imported", False)
                    and not candidate.get("index_missing", False)
                ],
            )
            pending_cover = self._nearest_covering_root_entry(
                entry.get("path", ""),
                actual_entries,
            )

            if is_indexed:
                entry["imported"] = True
                entry["index_missing"] = False
                entry["scope_only"] = False
                entry["covered_by"] = ""
                actual_entries.append(entry)
            elif indexed_cover is not None:
                entry["scope_only"] = True
                entry["covered_by"] = str(indexed_cover.get("path", ""))
                entry["imported"] = True
                entry["index_missing"] = False
                scope_entries.append(entry)
            elif not entry.get("imported", False) and pending_cover is not None:
                entry["scope_only"] = True
                entry["covered_by"] = str(pending_cover.get("path", ""))
                entry["imported"] = False
                entry["index_missing"] = True
                scope_entries.append(entry)
            else:
                entry["scope_only"] = False
                entry["covered_by"] = ""
                entry["index_missing"] = bool(entry.get("imported", False))
                actual_entries.append(entry)

        for entry in sorted(
            scope_entries,
            key=lambda candidate: len(
                normalized_root_path(candidate.get("path", ""))
            ),
        ):
            covering = self._nearest_covering_root_entry(
                entry.get("path", ""),
                actual_entries,
            )
            if covering is None:
                key = normalized_root_path(entry.get("path", ""))
                entry["scope_only"] = False
                entry["covered_by"] = ""
                entry["imported"] = key in indexed_keys
                entry["index_missing"] = False
                actual_entries.append(entry)
                continue
            usable = bool(
                covering.get("imported", False)
                and not covering.get("index_missing", False)
            )
            entry["scope_only"] = True
            entry["covered_by"] = str(covering.get("path", ""))
            entry["imported"] = usable
            entry["index_missing"] = not usable

    def _indexed_root_paths(self) -> set[str]:
        p = stats_path()
        if not p.exists():
            return set()
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
            roots = data.get("roots", []) if isinstance(data, dict) else []
            return {str(Path(x)) for x in roots if x}
        except Exception:
            return set()

    def _prune_saved_roots_against_index(self) -> None:
        """Reconcile saved roots without deleting valid scope aliases."""
        self._reconcile_scope_roots(self._indexed_root_paths())
        self._save_music_roots()

    def _save_music_roots(self) -> None:
        settings = self._load_settings()
        # Persist every folder the user added, including Pending folders.
        # The index files are database data, not the only copy of the UI list.
        settings["music_roots"] = [
            {
                "path": r["path"],
                "enabled": bool(r.get("enabled", True)),
                "imported": bool(r.get("imported", False)),
                "scope_only": bool(r.get("scope_only", False)),
                "covered_by": str(r.get("covered_by", "") or ""),
            }
            for r in self.music_roots if r.get("path")
        ]
        self._save_settings(settings)

    def _enabled_music_roots(self) -> list[Path]:
        return [Path(r["path"]) for r in self.music_roots if r.get("enabled", True)]

    def _pending_index_roots(self) -> list[Path]:
        """Return only real index owners that still require a scan."""
        return [
            Path(entry["path"])
            for entry in self.music_roots
            if (
                not entry.get("scope_only", False)
                and (
                    not entry.get("imported", False)
                    or entry.get("index_missing", False)
                )
            )
        ]

    def _collapse_overlapping_scan_roots(self, roots: list[Path]) -> list[Path]:
        """Keep shallowest roots so one scan never rereads a selected child."""
        unique: dict[str, Path] = {}
        for root in roots:
            key = normalized_root_path(root)
            if key:
                unique.setdefault(key, Path(root))
        collapsed: list[Path] = []
        for root in sorted(
            unique.values(),
            key=lambda value: len(normalized_root_path(value)),
        ):
            if any(root_contains(parent, root) for parent in collapsed):
                continue
            collapsed.append(root)
        return collapsed

    def _rescan_plan_for_paths(
        self,
        selected_paths: set[str],
    ) -> tuple[list[Path], dict[str, str]]:
        """Return scan targets plus the index owner for each target."""
        entries_by_key = {
            normalized_root_path(entry.get("path", "")): entry
            for entry in self.music_roots
            if entry.get("path")
        }
        target_owner_pairs: list[tuple[Path, Path]] = []
        for selected_path in selected_paths:
            entry = entries_by_key.get(normalized_root_path(selected_path))
            if entry is None:
                continue
            if entry.get("scope_only", False):
                covering = entries_by_key.get(
                    normalized_root_path(entry.get("covered_by", ""))
                )
                if covering is None:
                    covering = self._nearest_covering_root_entry(
                        entry.get("path", "")
                    )
                if covering is not None and covering.get("imported", False):
                    target_owner_pairs.append(
                        (
                            Path(str(entry.get("path", ""))),
                            Path(str(covering.get("path", ""))),
                        )
                    )
            elif entry.get("imported", False):
                path = Path(str(entry.get("path", "")))
                target_owner_pairs.append((path, path))

        targets = self._collapse_overlapping_scan_roots(
            [target for target, _owner in target_owner_pairs]
        )
        target_keys = {normalized_root_path(target) for target in targets}
        owners = {
            str(target): str(owner)
            for target, owner in target_owner_pairs
            if normalized_root_path(target) in target_keys
        }
        return targets, owners

    def _rescan_roots_for_paths(self, selected_paths: set[str]) -> list[Path]:
        """Compatibility helper used by tests and simple callers."""
        roots, _owners = self._rescan_plan_for_paths(selected_paths)
        return roots

    def on_music_root_item_changed(self, item: QListWidgetItem) -> None:
        raw = item.data(Qt.UserRole)
        if not raw:
            return
        for entry in self.music_roots:
            if entry.get("path") == str(raw):
                entry["enabled"] = item.checkState() == Qt.Checked
                break
        self._save_music_roots()
        self._update_music_roots_label()

    def _update_music_roots_label(self) -> None:
        enabled = sum(1 for r in self.music_roots if r.get("enabled", True))
        self.lbl_music_roots.setText(
            f"Enabled: {enabled} / Total: {len(self.music_roots)}"
        )

    # ---------- tiny helpers ----------
    def _is_exported_playlist(self, pl: Path) -> bool:
        stem = pl.stem.lower()
        return (
            stem.startswith("fixed_")
            or stem.startswith("draft_fixed_")
            or stem.endswith("_selected")
            or "_selected" in stem
        )

    def _show_import_hint_once(self) -> None:
        # Saved progress means this playlist is already fully or partially
        # repaired and can go straight to manual review. The onboarding hint is
        # useful only when every imported playlist is genuinely new.
        if not self.playlists or self._persisted_progress_keys:
            return

        settings = self._load_settings()
        if settings.get("hide_import_hint", False):
            return

        chk = QCheckBox("Don't show this again")

        box = QMessageBox(self)
        box.setIcon(QMessageBox.Information)
        box.setWindowTitle("Imported")
        box.setText(
            "These playlists do not have saved repair history yet. "
            "Run Repair (Safe) before manual review."
        )
        box.setCheckBox(chk)
        box.exec()

        if chk.isChecked():
            settings["hide_import_hint"] = True
            self._save_settings(settings)

    def _safe_str(self, v) -> str:
        try:
            return (str(v) if v is not None else "").strip()
        except Exception:
            return ""

    def _set_target_text(self, text: str) -> None:
        self.lbl_target.setText(text)
        # The label is deliberately allowed to clip instead of widening the
        # window.  Keep the complete value available on hover.
        self.lbl_target.setToolTip("" if text == "Target: (none)" else text)

    # ---------- UI helpers ----------
    def _set_busy(self, busy: bool, message: str | None = None):
        self._busy = busy

        buttons = [
            getattr(self, "btn_add_music", None),
            getattr(self, "btn_remove_music", None),
            getattr(self, "btn_clear_music", None),
            getattr(self, "btn_scan", None),
            getattr(self, "btn_rescan", None),
            getattr(self, "btn_import_pl", None),
            getattr(self, "btn_repair_safe", None),
            getattr(self, "btn_open_reports", None),
            getattr(self, "btn_browse_choice", None),
            getattr(self, "btn_apply_choice", None),
            getattr(self, "btn_save_fixed", None),
        ]

        for b in buttons:
            if b is None:
                continue
            try:
                b.setEnabled(not busy)
            except Exception:
                pass

        # The root checkboxes define the candidate scope used by Repair. Freeze
        # the whole list while any worker is running so the scope cannot mutate
        # halfway through a repair or scan.
        if getattr(self, "lst_music_roots", None) is not None:
            self.lst_music_roots.setEnabled(not busy)

        if getattr(self, "btn_open_reports", None) is not None:
            self.btn_open_reports.setEnabled(True)

        if getattr(self, "btn_cancel", None) is not None:
            self.btn_cancel.setEnabled(busy)

        if message is not None and getattr(self, "status_label", None) is not None:
            self.status_label.setText(message)

    def _setup_table(self, table: QTableWidget):
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setSelectionMode(QAbstractItemView.SingleSelection)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        # Pressing a row and brushing the top/bottom edge must not start the
        # QAbstractItemView drag-selection timer, which otherwise keeps
        # scrolling until it reaches the first/last row.
        table.setAutoScroll(False)

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setContentsMargins(8, 8, 8, 8)

        # =========================
        # Music roots panel
        # =========================
        roots_box = QGroupBox("Music Roots (checked = Repair scope)")
        roots_l = QVBoxLayout(roots_box)

        roots_btn_row = QHBoxLayout()
        self.btn_add_music = QPushButton("Add Music Folders")
        self.btn_remove_music = QPushButton("Remove Selected")
        self.btn_clear_music = QPushButton("Clear All")
        roots_btn_row.addWidget(self.btn_add_music, 2)
        roots_btn_row.addWidget(self.btn_remove_music, 1)
        roots_btn_row.addWidget(self.btn_clear_music, 1)
        roots_l.addLayout(roots_btn_row)

        self.lst_music_roots = QListWidget()
        self.lst_music_roots.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.lst_music_roots.itemChanged.connect(self.on_music_root_item_changed)
        roots_l.addWidget(self.lst_music_roots)

        self.lbl_music_roots = QLabel("Enabled: 0 / Imported: 0")
        self.lbl_music_roots.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        roots_l.addWidget(self.lbl_music_roots)

        # =========================
        # Fixed controls area (rows)
        # =========================
        row_scan = QHBoxLayout()
        self.btn_scan = QPushButton("Scan New Folders")
        self.btn_rescan = QPushButton("Rescan Selected")
        row_scan.addWidget(self.btn_scan)
        row_scan.addWidget(self.btn_rescan)

        row2 = QHBoxLayout()
        self.btn_import_pl = QPushButton("Import Playlist(s)")
        self.btn_repair_safe = QPushButton("Repair (Safe)")
        self.btn_open_reports = QPushButton("Open Reports Folder")
        row2.addWidget(self.btn_import_pl)
        row2.addWidget(self.btn_repair_safe)
        row2.addWidget(self.btn_open_reports)

        self.status_label = QLabel("Idle")
        # Status messages may contain a full folder path.  Do not let an
        # unbroken path become the minimum width of the whole main window.
        self.status_label.setMinimumWidth(0)
        self.status_label.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Preferred)
        self.scan_count_label = QLabel("Indexed: -")
        self.scan_count_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)

        status_row = QHBoxLayout()
        status_row.addWidget(self.status_label, 1)
        status_row.addWidget(self.scan_count_label, 0)

        self.progress = QProgressBar()
        self.progress.setRange(0, 100)
        self.btn_cancel = QPushButton("Cancel")
        self.btn_cancel.setEnabled(False)
        self.btn_cancel.setToolTip("Stop the current scan, repair, save, or folder removal safely.")

        progress_row = QHBoxLayout()
        progress_row.addWidget(self.progress, 1)
        progress_row.addWidget(self.btn_cancel, 0)

        search_row = QHBoxLayout()
        self.edt_search = QLineEdit()
        self.edt_search.setPlaceholderText("Search song / filename / path…")
        self.btn_clear_search = QPushButton("Clear")
        search_row.addWidget(QLabel("Search:"), 0)
        search_row.addWidget(self.edt_search, 1)
        search_row.addWidget(self.btn_clear_search, 0)

        controls = QHBoxLayout()
        self.lbl_target = QLabel("Target: (none)")
        # The selected target includes its full path and EXTINF text.  QLabel's
        # default minimum-size hint uses that entire string, which can force a
        # vertical splitter (and then the native window) wider than the screen.
        # Keep the complete text, but allow the layout to clip it horizontally.
        self.lbl_target.setMinimumWidth(0)
        self.lbl_target.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Preferred)

        self.cmb_view = QComboBox()
        self.cmb_view.addItems(["Unresolved", "Resolved"])
        self.cmb_view.setCurrentIndex(0)

        self.btn_browse_choice = QPushButton("Browse…")
        self.btn_apply_choice = QPushButton("Apply")
        self.btn_save_fixed = QPushButton("Save Fixed Playlist")

        controls.addWidget(self.lbl_target, 4)
        controls.addWidget(QLabel("View:"), 0)
        controls.addWidget(self.cmb_view, 0)
        controls.addWidget(self.btn_browse_choice, 1)
        controls.addWidget(self.btn_apply_choice, 1)
        controls.addWidget(self.btn_save_fixed, 2)

        # =========================
        # Unified repair queue + candidates
        # =========================
        boxA = QGroupBox("△ AMBIGUOUS / ✕ FAILED  (select row → choose candidate or Browse → Apply → Save)")
        boxA_l = QVBoxLayout(boxA)
        self.tbl_repair = QTableWidget(0, 5)
        self.tbl_repair.setHorizontalHeaderLabels(["Status", "Playlist", "EXTINF", "Original Path", "Notes"])
        self.tbl_repair.horizontalHeader().setStretchLastSection(True)
        # Column-header sorting was removed: changing sort keys made the repair
        # workflow order difficult to follow, especially after Apply.
        self.tbl_repair.horizontalHeader().setSectionsClickable(False)
        self.tbl_repair.setSortingEnabled(False)
        self.tbl_repair.setColumnWidth(0, 64)
        self._setup_table(self.tbl_repair)
        boxA_l.addWidget(self.tbl_repair)

        boxC = QGroupBox("Candidates / Picked file")
        boxC_l = QVBoxLayout(boxC)
        self.lst_candidates = QListWidget()
        boxC_l.addWidget(self.lst_candidates)

        inner_splitter = QSplitter(Qt.Vertical)
        inner_splitter.addWidget(boxA)
        inner_splitter.addWidget(boxC)
        inner_splitter.setStretchFactor(0, 5)
        inner_splitter.setStretchFactor(1, 2)

        # =========================
        # Lower container: fixed rows + inner splitter
        # =========================
        lower = QWidget()
        lower_layout = QVBoxLayout(lower)
        lower_layout.setContentsMargins(0, 0, 0, 0)
        lower_layout.addLayout(row_scan)
        lower_layout.addLayout(row2)
        lower_layout.addLayout(status_row)
        lower_layout.addLayout(progress_row)
        lower_layout.addLayout(search_row)
        lower_layout.addLayout(controls)
        lower_layout.addWidget(inner_splitter, 1)

        # =========================
        # Outer splitter: Roots vs Lower
        # =========================
        outer_splitter = QSplitter(Qt.Vertical)
        outer_splitter.addWidget(roots_box)
        outer_splitter.addWidget(lower)
        outer_splitter.setStretchFactor(0, 1)
        outer_splitter.setStretchFactor(1, 9)
        outer_splitter.setChildrenCollapsible(False)
        outer_splitter.setHandleWidth(8)
        # Resize the panes continuously with the handle.  Long labels are
        # shrinkable above, so live resizing cannot widen the native window.
        outer_splitter.setOpaqueResize(True)
        inner_splitter.setChildrenCollapsible(False)
        inner_splitter.setHandleWidth(8)
        inner_splitter.setOpaqueResize(True)

        # Both Music Roots and the repair/candidate areas can be resized by
        # dragging their splitter handles.
        outer_splitter.setSizes([160, 900])
        inner_splitter.setSizes([520, 260])

        layout.addWidget(outer_splitter, 1)

        # =========================
        # Info button in status bar (右下角)
        # =========================
        self.lbl_build = QLabel(f"v{APP_VERSION}")
        self.lbl_build.setToolTip("Build identifier. If this is not visible, an older copy is running.")
        self.statusBar().addPermanentWidget(self.lbl_build)

        self.btn_info = QPushButton("ⓘ")
        self.btn_info.setToolTip("About / Links")
        self.btn_info.setFixedWidth(32)
        self.statusBar().addPermanentWidget(self.btn_info)

        # =========================
        # Signals
        # =========================
        self.btn_add_music.clicked.connect(self.on_add_music)
        self.btn_remove_music.clicked.connect(self.on_remove_music_roots)
        self.btn_clear_music.clicked.connect(self.on_clear_music_roots)

        self.btn_scan.clicked.connect(self.on_scan_index)
        self.btn_rescan.clicked.connect(self.on_rescan_selected)
        self.btn_import_pl.clicked.connect(self.on_import_playlists)
        self.btn_repair_safe.clicked.connect(self.on_repair_safe)
        self.btn_open_reports.clicked.connect(self.on_open_reports)

        self.tbl_repair.itemSelectionChanged.connect(self.on_repair_selected)

        self.btn_apply_choice.clicked.connect(self.on_apply_choice)
        self.btn_browse_choice.clicked.connect(self.on_browse_choice)
        self.btn_save_fixed.clicked.connect(self.on_save_fixed)
        self.btn_cancel.clicked.connect(self.on_cancel_task)

        self.cmb_view.currentIndexChanged.connect(self.on_view_mode_changed)

        self.edt_search.textChanged.connect(self.on_search_changed)
        self.btn_clear_search.clicked.connect(lambda: self.edt_search.setText(""))

        self.btn_info.clicked.connect(self.on_about)

    def _refresh_music_roots_ui(self):
        self.lst_music_roots.blockSignals(True)
        self.lst_music_roots.clear()
        for entry in self.music_roots:
            path = str(entry.get("path", ""))
            available = Path(path).exists()
            imported = bool(entry.get("imported", False))
            index_missing = bool(entry.get("index_missing", False))
            waiting_for_scan = bool(entry.get("scope_only", False)) and not imported
            if not available:
                suffix = "  [Folder missing]"
            elif waiting_for_scan:
                suffix = "  [Needs scan]"
            elif index_missing:
                suffix = "  [Index missing]"
            elif not imported:
                suffix = "  [Needs scan]"
            else:
                suffix = "  [Ready]"
            it = QListWidgetItem(path + suffix)
            it.setData(Qt.UserRole, path)
            it.setFlags(it.flags() | Qt.ItemIsUserCheckable)
            it.setCheckState(Qt.Checked if entry.get("enabled", True) else Qt.Unchecked)
            if not available:
                it.setToolTip("This folder is currently unavailable. Reconnect the drive or restore the folder.")
            elif waiting_for_scan:
                it.setToolTip("This folder is waiting for an index. Use Scan New Folders.")
            elif index_missing:
                it.setToolTip("The folder exists, but its index record is missing. Scan or rescan this Music Root before Repair.")
            elif not imported:
                it.setToolTip("This folder has not been indexed yet. Use Scan New Folders.")
            else:
                it.setToolTip("This folder is ready to use for Repair.")
            self.lst_music_roots.addItem(it)
        self.lst_music_roots.blockSignals(False)
        self._update_music_roots_label()

    def _read_playlist_text(self, playlist: Path) -> str:
        """Read text playlists without silently turning a wrong encoding into mojibake."""
        raw = playlist.read_bytes()
        encodings = ("utf-8-sig", "utf-8", "utf-16", "cp932", "cp950", "big5", "cp1252")
        for encoding in encodings:
            try:
                return raw.decode(encoding)
            except (UnicodeDecodeError, LookupError):
                continue
        # Last resort: keep the file importable while clearly replacing only undecodable bytes.
        return raw.decode("utf-8", errors="replace")

    def _parse_xlsx_entries(self, playlist: Path) -> list[dict]:
        """Parse a Roon XLSX export for the pre-repair preview."""
        wb = load_workbook(playlist, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                return []
            header_map = {str(h).strip().lower(): i for i, h in enumerate(header) if h is not None}
            path_idx = header_map.get("path")
            if path_idx is None:
                raise ValueError("XLSX missing 'Path' column")

            # Use Roon metadata when present; otherwise fall back to the filename.
            title_idx = next((header_map[k] for k in ("title", "track title", "name") if k in header_map), None)
            artist_idx = next((header_map[k] for k in ("artist", "track artist", "artists") if k in header_map), None)
            entries: list[dict] = []
            for row in rows:
                if not row:
                    continue
                value = row[path_idx] if path_idx < len(row) else None
                original_path = str(value).strip() if value is not None else ""
                if not original_path:
                    continue
                title = str(row[title_idx]).strip() if title_idx is not None and title_idx < len(row) and row[title_idx] is not None else ""
                artist = str(row[artist_idx]).strip() if artist_idx is not None and artist_idx < len(row) and row[artist_idx] is not None else ""
                display = " - ".join(part for part in (artist, title) if part) or Path(original_path).stem
                entries.append({
                    "playlist": str(playlist),
                    "pl_key": self.runner.canonical_key(playlist),
                    "row_index": str(len(entries)),
                    "extinf_display": display,
                    "original_path": original_path,
                    "notes": "[NOT REPAIRED]",
                    "candidates": [],
                })
            return entries
        finally:
            wb.close()

    def _parse_playlist_entries(self, playlist: Path) -> list[dict]:
        if playlist.suffix.lower() == ".xlsx":
            return self._parse_xlsx_entries(playlist)

        text = self._read_playlist_text(playlist)
        lines = text.splitlines()
        entries: list[dict] = []
        pending_extinf = ""
        for line in lines:
            stripped = line.strip().lstrip("\ufeff")
            if not stripped:
                continue
            if stripped.upper().startswith("#EXTINF"):
                pending_extinf = stripped
                continue
            if stripped.startswith("#"):
                continue
            display = pending_extinf
            if "," in pending_extinf:
                display = pending_extinf.split(",", 1)[1].strip()
            entries.append({
                "playlist": str(playlist),
                "pl_key": self.runner.canonical_key(playlist),
                "row_index": str(len(entries)),
                "extinf_display": display or Path(stripped).stem,
                "original_path": stripped,
                "notes": "[NOT REPAIRED]",
                "candidates": [],
            })
            pending_extinf = ""
        return entries

    # ---------- view building ----------
    def _reload_reports_cache(self) -> None:
        """Load only active session reports or explicitly saved progress.

        Repair output is provisional until Save. A provisional report must not
        reappear after importing another playlist or restarting the app.
        """
        self._report_rows_by_key = {}
        self._persisted_progress_keys = set()
        self._persisted_artifacts_by_key = {}
        self.reports_path.mkdir(parents=True, exist_ok=True)

        for playlist_order, pl in enumerate(self.playlists):
            pl_key = self.runner.canonical_key(pl)
            report_csv = None
            if pl_key in self._session_repaired_keys:
                candidate = self._session_report_for(pl)
                if candidate.exists():
                    report_csv = candidate
            elif self._has_saved_progress(pl_key, pl):
                candidate = self._persisted_artifacts_by_key[pl_key]["report"]
                if candidate.exists():
                    report_csv = candidate
                    self._persisted_progress_keys.add(pl_key)

            if report_csv is None:
                continue
            rows = self.runner._read_report_rows(report_csv)
            if rows:
                # Path-only M3U and Roon XLSX reports do not contain EXTINF text.
                # Reuse the imported playlist preview by row index so Resolved still
                # identifies each song instead of showing an empty title column.
                try:
                    preview = {str(r.get("row_index", "")): r for r in self._parse_playlist_entries(pl)}
                except Exception:
                    preview = {}
                for rr in rows:
                    idx = str(rr.get("row_index", rr.get("_i", "")))
                    src = preview.get(idx, {})
                    if not self._safe_str(rr.get("extinf_display")):
                        rr["extinf_display"] = self._safe_str(src.get("extinf_display"))
                    if not self._safe_str(rr.get("original_path")):
                        rr["original_path"] = self._safe_str(src.get("original_path"))
                self._report_rows_by_key[pl_key] = rows

                # Saved manual choices are part of the repair result, not merely
                # a display detail. Load them into the export snapshot no matter
                # which view (Unresolved or Resolved) is currently open.
                if (
                    pl_key in self._persisted_progress_keys
                    and pl_key not in self._provisional_reset_keys
                ):
                    disk_selections = self._load_selections_for_key(pl_key)
                    if disk_selections:
                        current = self._selections_by_key.get(pl_key, {}) or {}
                        self._selections_by_key[pl_key] = {
                            **disk_selections,
                            **current,
                        }

    def _build_unresolved_rows(self) -> tuple[list[dict], list[dict]]:
        amb_all: list[dict] = []
        fail_all: list[dict] = []

        for pl in self.playlists:
            pl_key = self.runner.canonical_key(pl)

            is_exported = self._is_exported_playlist(pl)
            has_saved_progress = pl_key in self._persisted_progress_keys

            report_rows = self._report_rows_by_key.get(pl_key, [])

            # Before Repair, always show the imported file itself when there is no usable
            # report in this session. This also covers fixed_*_selected.m3u8 files imported
            # as a new input instead of treating them as invisible historical progress.
            if pl_key not in self._session_repaired_keys and not has_saved_progress:
                pending_rows = self._parse_playlist_entries(pl)
                for r in pending_rows:
                    r["_bucket"] = "PENDING"
                amb_all.extend(pending_rows)
                continue

            if not report_rows:
                continue

            # --- selections sources ---
            # disk_sel: 以前 Save 過、寫在磁碟上的 selections（永遠視為已完成）
            # mem_sel : 本次 session Apply 但尚未 Save 的 selections
            disk_sel: dict[str, str] = {}
            if has_saved_progress and pl_key not in self._provisional_reset_keys:
                disk_sel = self._load_selections_for_key(pl_key) or {}

            mem_sel: dict[str, str] = self._selections_by_key.get(pl_key, {}) or {}

            # merged view for lookup
            merged_sel = {**disk_sel, **mem_sel}

            amb, fail = self.runner._classify_for_ui(report_rows, pl)

            # ✅ A) disk_sel：永遠 hide（因為它代表「上次已存檔」）
            if disk_sel:
                amb = [r for r in amb if str(r.get("row_index")) not in disk_sel]
                fail = [r for r in fail if str(r.get("row_index")) not in disk_sel]

            # B) In-memory selections can contain both already-saved choices and
            # brand-new unsaved edits.  Do not use the playlist-level _saved_keys
            # flag here: once a playlist has been saved once, that flag remains set
            # and would incorrectly hide every later unsaved ✓ when the user switches
            # views.  A row is unsaved only while its exact (playlist, row) key is in
            # _dirty_selection_ids and the Apply was made in Unresolved.
            if mem_sel:
                dirty_unresolved = {
                    row_id
                    for key, row_id in self._dirty_selection_ids
                    if key == pl_key
                    and self._selection_origin.get((key, row_id)) == "UNRESOLVED"
                }

                # Hide committed selections and edits made in Resolved; keep only
                # current unsaved Unresolved edits in this workspace.
                amb = [
                    r for r in amb
                    if str(r.get("row_index")) not in mem_sel
                    or str(r.get("row_index")) in dirty_unresolved
                ]
                fail = [
                    r for r in fail
                    if str(r.get("row_index")) not in mem_sel
                    or str(r.get("row_index")) in dirty_unresolved
                ]

                for rr in amb:
                    k = str(rr.get("row_index"))
                    if k in dirty_unresolved:
                        rr["notes"] = f"[SELECTED] {mem_sel[k]}"
                for rr in fail:
                    k = str(rr.get("row_index"))
                    if k in dirty_unresolved:
                        rr["notes"] = f"[RESCUED] {mem_sel[k]}"

            # ✅ 把 merged 保存回記憶體（避免你匯入 exported 後，disk_sel 覆蓋掉本次 Apply 的 mem_sel）
            #    但注意：原始歌單你原本的設計是「不吃 disk」，所以這裡只有 exported 才會 merge disk
            if has_saved_progress and merged_sel:
                self._selections_by_key[pl_key] = merged_sel

            amb_all.extend(amb)
            fail_all.extend(fail)

        return amb_all, fail_all

    def _build_resolved_rows(self) -> tuple[list[dict], list[dict]]:
        """
        Show resolved rows (Auto + Manual).
        Notes column will show [AUTO]/[MANUAL] after_path + status.
        """
        amb_rows: list[dict] = []
        fail_rows: list[dict] = []

        # local status sets (aligned with runner logic)
        AMBIG_STATUSES = {"AMBIGUOUS", "MULTI_MATCH", "MULTIPLE_MATCH", "CONFLICT", "DUPLICATE"}
        FAIL_STATUSES = {"FAILED", "NOT_FOUND", "MISSING", "ERROR"}
        RESOLVED_STATUSES = {"KEPT", "REPAIRED", "FIXED", "OK", "DONE", "SUCCESS", "RESOLVED"}

        def pick_written_path(rr: dict) -> str:
            # strict whitelist: only these keys can be treated as "final written path"
            FINAL_KEYS = (
                "written_path", "written",
                "final_path", "final",
                "resolved_path", "resolved",
                "picked_path", "picked",
                "chosen_path", "chosen",
                "selected_path", "selected",
                "output_path", "output",
                "result_path", "result",
                "target_path", "target",
                "matched_path", "matched",
            )
            for k in FINAL_KEYS:
                v = rr.get(k)
                s = self._safe_str(v)
                if s:
                    return s
            return ""

        for playlist_order, pl in enumerate(self.playlists):
            pl_key = self.runner.canonical_key(pl)

            has_saved_progress = pl_key in self._persisted_progress_keys
            saved_this_session = pl_key in self._saved_keys
            has_committed_progress = (has_saved_progress or saved_this_session) and pl_key not in self._provisional_reset_keys

            # Show resolved rows for this session's Repair, for progress loaded
            # from an exported fixed playlist, or immediately after Save in the
            # current session.  _reload_reports_cache() intentionally does not
            # treat the original source playlist as persisted progress, so the
            # separate saved_this_session flag is required here.
            if pl_key not in self._session_repaired_keys and not has_committed_progress:
                continue

            disk_sel = self._load_selections_for_key(pl_key) if has_committed_progress else {}
            mem_sel_all = self._selections_by_key.get(pl_key, {}) or {}
            # After Save, every in-memory choice is committed and must appear in
            # Resolved immediately, including rows fixed from Unresolved. Before
            # Save, preview only edits made while already auditing Resolved.
            if saved_this_session:
                mem_sel = dict(mem_sel_all)
            else:
                mem_sel = {
                    row_id: chosen
                    for row_id, chosen in mem_sel_all.items()
                    if self._selection_origin.get((pl_key, str(row_id))) == "RESOLVED"
                }
            selections = {**disk_sel, **mem_sel}

            report_rows = self._report_rows_by_key.get(pl_key, [])
            if not report_rows:
                continue

            for rr in report_rows:
                status = self._safe_str(rr.get("status")).upper()
                row_index = self._safe_str(rr.get("row_index", rr.get("_i", "")))
                extinf_display = self._safe_str(rr.get("extinf_display") or rr.get("extinf") or "")
                notes_raw = self._safe_str(rr.get("notes") or "")

                orig = self._safe_str(rr.get("original_path") or rr.get("original") or "")
                written = pick_written_path(rr)

                manual = bool(row_index and row_index in selections)
                after = selections.get(row_index, "").strip() if manual else (written or orig)

                # resolved 판단: manual exists OR report itself says resolved OR has a written/picked path
                auto_resolved = any(k in status for k in RESOLVED_STATUSES)
                is_resolved = manual or auto_resolved
                if not is_resolved:
                    continue

                # Parse candidates once. Resolved previously parsed the same
                # Notes field twice per row, which noticeably slowed large lists.
                try:
                    cands = self.runner._parse_candidates_from_notes(notes_raw) or []
                except Exception:
                    cands = []

                # Decide bucket (used internally for candidate handling only).
                if status in AMBIG_STATUSES:
                    bucket = "AMBIGUOUS"
                elif status in FAIL_STATUSES:
                    bucket = "FAILED"
                else:
                    bucket = "AMBIGUOUS" if len(cands) >= 2 else "FAILED"

                source_tag = "[MANUAL]" if manual else "[AUTO]"
                status_tag = f"(status={status})" if status else ""

                try:
                    row_order = int(float(row_index))
                except Exception:
                    row_order = 10**12

                row = {
                    "playlist": str(pl),
                    "pl_key": pl_key,
                    "row_index": row_index if row_index else self._safe_str(rr.get("_i", "")),
                    "extinf_display": extinf_display,
                    "original_path": orig,
                    "notes": f"{source_tag} {after} {status_tag}".strip(),
                    "candidates": cands,
                    "_bucket": bucket,
                    "_playlist_order": playlist_order,
                    "_row_order": row_order,
                }

                if bucket == "AMBIGUOUS":
                    amb_rows.append(row)
                else:
                    fail_rows.append(row)

        return amb_rows, fail_rows

    def _refresh_tables_from_mode(self) -> None:
        """Rebuild visible rows based on current view mode and cached report rows."""
        # clear active selection & candidates
        self.tbl_repair.clearSelection()
        self.lst_candidates.clear()
        self._active_target = None
        self._active_pl_key = None
        self._active_row_id = None
        self._set_target_text("Target: (none)")

        if self._view_mode == "RESOLVED":
            amb, fail = self._build_resolved_rows()
        else:
            amb, fail = self._build_unresolved_rows()

        # master (unfiltered)
        self._ambiguous_rows_all = amb
        self._failed_rows_all = fail

        # apply search filter (will also fill tables)
        self._apply_search_filter()

        # tiny status hint
        if self._view_mode == "RESOLVED":
            self.status_label.setText("View: Resolved (audit / fix wrong picks)")
        else:
            self.status_label.setText("View: Unresolved (needs action)")

    # ---------- actions ----------
    def on_view_mode_changed(self, idx: int):
        self._view_mode = "RESOLVED" if idx == 1 else "UNRESOLVED"
        self._refresh_tables_from_mode()

    def _has_unsaved_work(self) -> bool:
        if self._dirty_selection_ids:
            return True
        return bool(self._session_repaired_keys - self._saved_keys)

    def _confirm_discard_unsaved(self, action: str) -> bool:
        if not self._has_unsaved_work():
            return True
        answer = QMessageBox.question(
            self,
            "Unsaved repair changes",
            "The current repair result or manual changes have not been saved.\n\n"
            f"{action} will discard them from this session.\n\nContinue?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        return answer == QMessageBox.Yes

    def on_add_music(self):
        # Qt's native Windows folder picker only supports one folder.  Use the
        # Qt dialog and enable extended selection in both directory views so
        # the user can choose several sibling folders in one operation.
        dialog = QFileDialog(self, "Select Music Folders")
        dialog.setFileMode(QFileDialog.Directory)
        dialog.setOption(QFileDialog.ShowDirsOnly, True)
        dialog.setOption(QFileDialog.DontUseNativeDialog, True)

        # PySide6 QObject.findChildren() accepts one Qt type at a time;
        # passing a Python tuple raises TypeError. Configure both view types
        # separately so the non-native folder dialog supports multi-selection.
        for view_type in (QListView, QTreeView):
            for view in dialog.findChildren(view_type):
                view.setSelectionMode(QAbstractItemView.ExtendedSelection)

        if dialog.exec() != QDialog.Accepted:
            return

        folders = dialog.selectedFiles()
        if not folders:
            return

        # QFileDialog in non-native Directory mode may include the directory
        # currently being viewed in selectedFiles(), even when the user only
        # selected its child folders.  That caused the parent folder to be
        # added unexpectedly.  When multiple folders were selected, discard
        # the dialog's current directory and keep only the actual selections.
        current_dir = os.path.normcase(
            os.path.normpath(dialog.directory().absolutePath())
        )
        normalized_folders = []
        seen_folders = set()
        for folder in folders:
            normalized = os.path.normcase(os.path.normpath(str(folder)))
            if len(folders) > 1 and normalized == current_dir:
                continue
            if normalized in seen_folders or not Path(folder).is_dir():
                continue
            seen_folders.add(normalized)
            normalized_folders.append(str(Path(folder)))

        folders = sorted(
            normalized_folders,
            key=lambda value: len(normalized_root_path(value)),
        )
        if not folders:
            return

        existing = {
            normalized_root_path(r.get("path", ""))
            for r in self.music_roots
            if r.get("path")
        }
        added = 0
        duplicates = 0
        scope_added = 0

        for folder in folders:
            p = Path(folder)
            key = normalized_root_path(p)
            if key in existing:
                duplicates += 1
                continue

            entry = self._new_music_root_entry(p)
            self.music_roots.append(entry)
            existing.add(key)
            added += 1
            if entry.get("scope_only", False):
                scope_added += 1

        if added:
            # Adding a new parent above Pending child roots makes those children
            # scope aliases immediately. Imported child roots remain independent
            # until the new parent scan succeeds, so the working index is never
            # discarded early.
            self._reconcile_scope_roots(self._indexed_root_paths())
            # Save immediately so the list survives a restart even before Scan.
            self._save_music_roots()
            self._refresh_music_roots_ui()

        status_parts = []
        if added:
            pending_added = sum(
                1
                for entry in self.music_roots
                if (
                    normalized_root_path(entry.get("path", ""))
                    in {
                        normalized_root_path(folder)
                        for folder in folders
                    }
                    and not entry.get("scope_only", False)
                    and not entry.get("imported", False)
                )
            )
            if pending_added:
                status_parts.append(f"Needs scan: {pending_added} folder(s)")
            if scope_added:
                status_parts.append(f"Ready: {scope_added} folder(s)")
        if duplicates:
            status_parts.append(f"already added: {duplicates}")
        if status_parts:
            self.status_label.setText("; ".join(status_parts))
        else:
            self.status_label.setText("No folder was added")

    def on_remove_music_roots(self):
        if self._busy:
            QMessageBox.information(self, "Busy", "A task is already running. Please wait.")
            return
        items = self.lst_music_roots.selectedItems()
        if not items:
            return

        to_remove = set()
        for it in items:
            raw = it.data(Qt.UserRole)
            if raw:
                to_remove.add(str(raw))

        if not to_remove:
            return

        selected_entries = [
            entry
            for entry in self.music_roots
            if str(entry.get("path", "")) in to_remove
        ]
        scope_paths = {
            str(entry.get("path", ""))
            for entry in selected_entries
            if entry.get("scope_only", False)
        }
        index_root_paths = {
            str(entry.get("path", ""))
            for entry in selected_entries
            if not entry.get("scope_only", False)
        }

        if scope_paths:
            self.music_roots = [
                entry
                for entry in self.music_roots
                if str(entry.get("path", "")) not in scope_paths
            ]
            self._save_music_roots()
            self._refresh_music_roots_ui()

        if not index_root_paths:
            self.status_label.setText(
                f"Removed {len(scope_paths)} Repair scope folder(s); music index unchanged"
            )
            return

        # A retained child already has records inside the parent's index. Hand
        # those rows to the child when the parent is removed instead of forcing
        # an unnecessary disk rescan.
        preserve_paths = [
            Path(str(entry.get("path", "")))
            for entry in self.music_roots
            if (
                entry.get("path")
                and str(entry.get("path", "")) not in to_remove
                and any(
                    root_contains(parent, entry.get("path", ""), include_same=False)
                    for parent in index_root_paths
                )
            )
        ]

        self._pending_remove_paths = set(index_root_paths)
        self._last_action = "REMOVE_ROOTS"
        self._run_task(
            self.runner.remove_roots_from_index,
            roots_to_remove=[Path(p) for p in sorted(index_root_paths)],
            out_index=self.index_path,
            preserve_roots=preserve_paths,
        )

    def on_clear_music_roots(self):
        if self._busy:
            QMessageBox.information(self, "Busy", "A task is already running. Please wait.")
            return
        if not self.music_roots:
            return
        answer = QMessageBox.question(
            self,
            "Clear all Music Roots?",
            "This removes every Music Root from the list and deletes their entries "
            "from the music index.\n\nYour actual audio files will not be deleted.\n\nContinue?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if answer != QMessageBox.Yes:
            return
        roots_to_remove = [
            Path(r.get("path", ""))
            for r in self.music_roots
            if r.get("path") and not r.get("scope_only", False)
        ]
        if not roots_to_remove:
            self.music_roots = []
            self._save_music_roots()
            self._refresh_music_roots_ui()
            self.status_label.setText("Cleared all Repair scope folders; music index unchanged")
            return
        self._pending_remove_paths = {str(path) for path in roots_to_remove}
        self._last_action = "CLEAR_ROOTS"
        self._run_task(
            self.runner.remove_roots_from_index,
            roots_to_remove=roots_to_remove,
            out_index=self.index_path,
        )

    def on_scan_index(self):
        """Import only folders that have not yet been indexed.

        Checkboxes remain Repair-scope controls. Existing imported roots are
        deliberately skipped so adding one new folder never rescans a large
        established library.
        """
        if self._busy:
            QMessageBox.information(self, "Busy", "A task is already running. Please wait.")
            return

        scan_roots = self._pending_index_roots()
        if not scan_roots:
            QMessageBox.information(
                self,
                "No new folders",
                "There are no folders that need scanning.\n\n"
                "Use Rescan Selected only when you intentionally want to refresh an existing folder.",
            )
            return

        missing = [str(p) for p in scan_roots if not p.exists()]
        if missing:
            QMessageBox.warning(self, "Unavailable folder", "These new folders are unavailable:\n" + "\n".join(missing))
            return

        self.status_label.setText(f"Scanning {len(scan_roots)} new folder(s)…")
        self._last_action = "SCAN"
        self._run_task(self.runner.scan_index, music_roots=scan_roots, out_index=self.index_path)

    def on_rescan_selected(self):
        """Explicitly refresh imported roots selected by row highlight.

        Row selection controls this operation; checkbox state still controls
        Repair scope and is intentionally ignored here.
        """
        if self._busy:
            QMessageBox.information(self, "Busy", "A task is already running. Please wait.")
            return

        selected_paths = {
            str(item.data(Qt.UserRole))
            for item in self.lst_music_roots.selectedItems()
            if item.data(Qt.UserRole)
        }
        scan_roots, root_owners = self._rescan_plan_for_paths(selected_paths)
        if not scan_roots:
            QMessageBox.information(
                self,
                "Select imported folder",
                "Highlight one or more already imported folders in the list, then choose Rescan Selected.\n\n"
                "Scope-only folders update only their part of the covering index. "
                "The checkbox is only for Repair scope.",
            )
            return

        missing = [str(p) for p in scan_roots if not p.exists()]
        if missing:
            QMessageBox.warning(self, "Unavailable folder", "These selected folders are unavailable:\n" + "\n".join(missing))
            return

        names = "\n".join(str(p) for p in scan_roots)
        answer = QMessageBox.question(
            self,
            "Rescan selected folders?",
            "This will reread and replace the existing index entries for:\n\n"
            + names
            + "\n\nOther imported folders will remain unchanged.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if answer != QMessageBox.Yes:
            return

        self.status_label.setText(f"Rescanning {len(scan_roots)} selected folder(s)…")
        self._last_action = "SCAN"
        self._run_task(
            self.runner.scan_index,
            music_roots=scan_roots,
            out_index=self.index_path,
            root_owners=root_owners,
        )

    def on_import_playlists(self):
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select playlist(s)",
            "",
            "Playlists (*.m3u *.m3u8 *.xlsx);;M3U playlists (*.m3u *.m3u8);;Roon Excel exports (*.xlsx);;All files (*.*)"
        )
        if not files:
            return
        if not self._confirm_discard_unsaved("Importing another playlist"):
            return

        self.playlists = [Path(p) for p in files]
        self.status_label.setText(f"Loaded playlists: {len(self.playlists)}")

        # load reports cache then refresh view
        self._selections_by_key = {}          # ✅ 清掉上一輪 Apply 的記憶體選擇
        self._dirty_selection_ids.clear()      # 新 session 尚無未儲存修改
        self._session_repaired_keys = set()   # ✅ 新 session
        self._saved_keys = set()   # ✅ 新 session，尚未 Save
        self._provisional_reset_keys = set()
        self._reload_reports_cache()
        self._refresh_tables_from_mode()
        self._show_import_hint_once()

    def _begin_clean_provisional_rerun(self, current_keys: set[str]) -> None:
        """Reset current-session state while preserving previously saved files on disk."""
        self._provisional_reset_keys.update(current_keys)
        self._saved_keys.difference_update(current_keys)
        self._persisted_progress_keys.difference_update(current_keys)
        for pl_key in current_keys:
            self._selections_by_key.pop(pl_key, None)
        self._dirty_selection_ids = {
            key for key in self._dirty_selection_ids if key[0] not in current_keys
        }
        self._selection_origin = {
            key: origin
            for key, origin in self._selection_origin.items()
            if key[0] not in current_keys
        }

    def _saved_source_playlist(self, playlist: Path) -> Path | None:
        """Return the original source recorded for an imported fixed playlist."""
        pl_key = self.runner.canonical_key(playlist)
        artifacts = (
            self._persisted_artifacts_by_key.get(pl_key)
            or self._saved_progress_artifacts(pl_key, playlist)
        )
        if artifacts is None:
            return None
        marker = artifacts["progress"]
        try:
            data = json.loads(marker.read_text(encoding="utf-8"))
            source = data.get("source_playlist")
            if not source:
                return None
            source_path = Path(source)
            return source_path if source_path.exists() else None
        except Exception:
            return None

    def on_repair_safe(self):
        if self._busy:
            QMessageBox.information(self, "Busy", "A task is already running. Please wait.")
            return
        if not self.playlists:
            QMessageBox.warning(self, "No playlist", "Please import playlist(s) first.")
            return
        enabled_roots = self._enabled_music_roots()
        if not enabled_roots:
            QMessageBox.warning(self, "No music root", "Please enable at least one music folder before Repair.")
            return

        missing_index_roots = [
            str(r.get("path", ""))
            for r in self.music_roots
            if r.get("enabled", True) and r.get("index_missing", False)
        ]
        if missing_index_roots:
            QMessageBox.warning(
                self,
                "Music Root index missing",
                "These checked Music Roots exist, but their index records are missing:\n\n"
                + "\n".join(missing_index_roots)
                + "\n\nUse Scan New Folders or select them and use Rescan Selected before Repair.",
            )
            return

        if not self.index_path.exists():
            QMessageBox.warning(self, "No index", f"Index not found: {self.index_path}\nPlease use Scan New Folders first.")
            return

        current_keys = {self.runner.canonical_key(pl) for pl in self.playlists}
        is_repeat_repair = bool(current_keys & self._session_repaired_keys)

        if is_repeat_repair:
            answer = QMessageBox.question(
                self,
                "Run Repair again?",
                "Repair has already been run for the current playlist.\n\n"
                "Running it again will discard all unsaved Repair results and manual changes "
                "marked with ✓, then repair the playlist again using the currently checked "
                "Music Roots.\n\n"
                "Saved progress and previously exported playlists will not be deleted.\n\n"
                "Run Repair again?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if answer != QMessageBox.Yes:
                return

            # Start a genuinely clean provisional view. Saved files remain on
            # disk, but old selections and old Resolved rows are ignored until
            # the new provisional result is explicitly saved.
            self._begin_clean_provisional_rerun(current_keys)

        # 防呆：避免誤按 Repair 覆寫 report
        has_any_report = any(
            self._has_saved_progress(self.runner.canonical_key(pl), pl)
            for pl in self.playlists
        )

        if has_any_report and not is_repeat_repair:
            box = QMessageBox(self)
            box.setIcon(QMessageBox.Question)
            box.setWindowTitle("Repair report exists")
            box.setText(
                "Saved repair progress already exists.\n\n"
                "Resume keeps that saved progress. Re-run creates a provisional new result; "
                "the saved progress is not replaced unless you press Save.\n\n"
                "What do you want to do?"
            )

            btn_resume = box.addButton("Resume (load existing report)", QMessageBox.AcceptRole)
            btn_rerun = box.addButton("Re-run Repair (provisional)", QMessageBox.DestructiveRole)
            btn_cancel = box.addButton("Cancel", QMessageBox.RejectRole)

            box.exec()
            clicked = box.clickedButton()

            if clicked == btn_resume:
                self._session_repaired_keys = set()
                self._provisional_reset_keys.difference_update(current_keys)
                self._reload_reports_cache()
                self._refresh_tables_from_mode()
                self.status_label.setText("Resumed from saved progress.")
                return
            if clicked == btn_cancel:
                return
            # clicked == btn_rerun -> create a clean provisional session. If the
            # imported file is a previously exported fixed playlist, re-run from
            # its recorded original source so the behavior matches re-running in
            # the same session.
            original_playlists = []
            for pl in self.playlists:
                original_playlists.append(self._saved_source_playlist(pl) or pl)
            self.playlists = original_playlists
            current_keys = {self.runner.canonical_key(pl) for pl in self.playlists}
            self._begin_clean_provisional_rerun(current_keys)

        # clear UI before repair
        self._ambiguous_rows = []
        self._failed_rows = []
        self._amb_by_id = {}
        self._fail_by_id = {}
        self._active_target = None
        self._active_pl_key = None
        self._active_row_id = None
        self._set_target_text("Target: (none)")
        self._fill_table(self.tbl_repair, [])
        self.lst_candidates.clear()
        self._session_repaired_keys = {self.runner.canonical_key(pl) for pl in self.playlists}
        self._pending_repair_keys = set(self._session_repaired_keys)
        self._last_action = "REPAIR"

        self._run_task(
            self.runner.repair_playlists,
            playlists=self.playlists,
            index_path=self.index_path,
            out_dir=self.reports_path,
            mode="safe",
            dry_run=False,
            enabled_roots=enabled_roots,
            session_reports=True,
        )

    def on_open_reports(self):
        try:
            self.reports_path.mkdir(parents=True, exist_ok=True)
            os.startfile(str(self.reports_path))
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def on_repair_selected(self):
        if not (self.tbl_repair.selectionModel() and self.tbl_repair.selectionModel().hasSelection()):
            return
        self._active_pl_key, self._active_row_id = self._selected_row_id(self.tbl_repair)
        vis_row = self._selected_visual_row(self.tbl_repair)
        self._active_target = None
        if vis_row is not None:
            item = self.tbl_repair.item(vis_row, 0)
            data = item.data(Qt.UserRole) if item else None
            if isinstance(data, dict):
                self._active_target = str(data.get("bucket") or "") or None
        self._refresh_candidates_panel()

    def on_browse_choice(self):
        if self._active_target is None or self._active_row_id is None or self._active_pl_key is None:
            QMessageBox.warning(self, "No selection", "Please select a row from the repair list first.")
            return

        file, _ = QFileDialog.getOpenFileName(
            self,
            "Pick the correct audio file",
            "",
            "Audio files (*.mp3 *.aac *.ogg *.opus *.mp4 *.m4a *.flac *.alac *.wav *.aif *.aiff *.aifc *.ape *.wv *.dsf *.dff);;All files (*.*)",
        )
        if not file:
            return

        self.lst_candidates.clear()
        item = QListWidgetItem(file)
        item.setData(Qt.UserRole, file)
        self.lst_candidates.addItem(item)
        self.lst_candidates.setCurrentRow(0)

    def on_apply_choice(self):
        if self._active_target is None or self._active_row_id is None or self._active_pl_key is None:
            QMessageBox.warning(self, "No selection", "Please select a row from the repair list first.")
            return

        chosen = self._current_candidate()
        if not chosen:
            QMessageBox.warning(self, "No file", "Please select a candidate (or Browse…) first.")
            return

        pl_key = self._active_pl_key
        row_id = self._active_row_id

        # only in-memory, do NOT persist to disk here
        self._selections_by_key.setdefault(pl_key, {})[row_id] = chosen
        self._dirty_selection_ids.add((pl_key, row_id))
        self._selection_origin[(pl_key, row_id)] = self._view_mode

        tag = "[SELECTED]" if self._active_target == "AMBIGUOUS" else "[RESCUED]"
        table = self.tbl_repair

        vis_row = self._selected_visual_row(table)
        if vis_row is not None:
            # Apply changes only this row in place. No re-sort or row movement.
            status_item = QTableWidgetItem("✓")
            status_item.setTextAlignment(Qt.AlignCenter)
            status_item.setData(Qt.UserRole, {"bucket": self._active_target})
            table.setItem(vis_row, 0, status_item)
            if self._view_mode == "RESOLVED":
                table.setItem(vis_row, 4, QTableWidgetItem(f"[MANUAL] {chosen}"))
            else:
                table.setItem(vis_row, 4, QTableWidgetItem(f"{tag} {chosen}"))

        self.status_label.setText(f"Applied (not saved): key={pl_key} row={row_id}")

        # If in unresolved view, you might want to keep the row visible until Save.
        # If in resolved view, also keep visible to allow further audit.
        # (No auto-removal here; removal happens after Save+reload.)

    def on_save_fixed(self):
        """
        Save/export final playlists.
        This is the ONLY step that writes fixed_*.m3u or .m3u8
        AND the ONLY step that persists selections_*.json
        """
        if self._busy:
            QMessageBox.information(self, "Busy", "A task is already running. Please wait.")
            return
        if not self.playlists:
            QMessageBox.warning(self, "No playlist", "Please import playlist(s) first.")
            return

        format_label, accepted = QInputDialog.getItem(
            self,
            "Playlist format",
            "Save as:",
            ["M3U8 (UTF-8, recommended)", "M3U (UTF-8 with BOM, legacy compatibility)"],
            0,
            False,
        )
        if not accepted:
            return
        export_extension = ".m3u" if format_label.startswith("M3U (") else ".m3u8"

        jobs = []
        pending_keys: list[str] = []
        pending_snapshots: list[dict] = []

        for pl in self.playlists:
            pl_key = self.runner.canonical_key(pl)
            if pl_key in self._session_repaired_keys:
                report_csv = self._session_report_for(pl)
            elif self._has_saved_progress(pl_key, pl):
                report_csv = self._persisted_artifacts_by_key[pl_key]["report"]
            else:
                continue
            if not report_csv.exists():
                continue

            out_m3u = self.runner.export_path_for(self.reports_path, pl, export_extension)
            # Take a detached snapshot so the worker receives exactly the choices
            # visible at the moment Save is pressed.  Also attach original-path
            # fallbacks: this protects manual repairs if a report row id is read
            # with a different textual form (for example 1 vs 1.0).
            selections = dict(self._selections_by_key.get(pl_key, {}) or {})
            rows_by_id = {
                str(r.get("row_index", r.get("_i", ""))).strip(): r
                for r in (self._report_rows_by_key.get(pl_key, []) or [])
            }
            selection_records = []
            for selected_row_id, selected_path in selections.items():
                rr = rows_by_id.get(str(selected_row_id).strip(), {})
                selection_records.append({
                    "row_index": str(selected_row_id).strip(),
                    "original_path": self._safe_str(rr.get("original_path") or rr.get("original") or ""),
                    "chosen_path": str(selected_path),
                })

            jobs.append({
                "report_csv": str(report_csv),
                "out_m3u": str(out_m3u),
                "selections": selections,
                "selection_records": selection_records,
                "index_path": str(self.index_path),
                "source_playlist": str(pl),
            })
            pending_keys.append(pl_key)
            pending_snapshots.append({
                "source_key": pl_key,
                "source_playlist": pl,
                "original_source": self._saved_source_playlist(pl) or pl,
                "report_csv": report_csv,
                "saved_output": out_m3u,
                "selections": selections,
            })

        if not jobs:
            QMessageBox.critical(self, "Missing report", "No active Repair result or saved progress was found.\nRun Repair (Safe) first.")
            return

        output_counts: dict[str, int] = {}
        output_names: dict[str, str] = {}
        for job in jobs:
            output_path = Path(job["out_m3u"])
            normalized = self._normalized_playlist_path(output_path)
            output_counts[normalized] = output_counts.get(normalized, 0) + 1
            output_names[normalized] = output_path.name
        duplicate_outputs = sorted(
            output_names[path]
            for path, count in output_counts.items()
            if count > 1
        )
        if duplicate_outputs:
            QMessageBox.critical(
                self,
                "Duplicate output filename",
                "Two or more imported playlists would be saved to the same file:\n\n"
                + "\n".join(duplicate_outputs)
                + "\n\nRename one of the source playlists so each fixed playlist "
                "has a different filename.",
            )
            return

        existing_outputs = sorted({
            Path(job["out_m3u"]).name
            for job in jobs
            if Path(job["out_m3u"]).exists()
        })
        if existing_outputs:
            preview = "\n".join(existing_outputs[:10])
            if len(existing_outputs) > 10:
                preview += f"\n…and {len(existing_outputs) - 10} more"
            answer = QMessageBox.question(
                self,
                "Replace existing fixed playlist?",
                "These output files already exist:\n\n"
                + preview
                + "\n\nReplace them with the current repair result?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if answer != QMessageBox.Yes:
                return

        self._pending_save_keys = pending_keys
        self._pending_save_snapshots = pending_snapshots
        self._last_action = "SAVE"
        self._run_task(self.runner.export_fixed_multi, jobs=jobs)
    
    def on_about(self):
        AboutDialog(Path(self.app_data), self).exec()

    # ---------- internals ----------
    def _rebuild_row_maps(self):
        self._amb_by_id = {}
        self._fail_by_id = {}

        for r in self._ambiguous_rows:
            pl_key = str(r.get("pl_key", "")).strip()
            row_id = str(r.get("row_index", "")).strip()
            if pl_key and row_id:
                self._amb_by_id[f"{pl_key}::{row_id}"] = r

        for r in self._failed_rows:
            pl_key = str(r.get("pl_key", "")).strip()
            row_id = str(r.get("row_index", "")).strip()
            if pl_key and row_id:
                self._fail_by_id[f"{pl_key}::{row_id}"] = r

    def _refresh_candidates_panel(self):
        self.lst_candidates.clear()

        if self._active_target is None or self._active_row_id is None or self._active_pl_key is None:
            self._set_target_text("Target: (none)")
            return

        key = f"{self._active_pl_key}::{self._active_row_id}"
        r = self._amb_by_id.get(key) if self._active_target == "AMBIGUOUS" else self._fail_by_id.get(key)

        if not r:
            self._set_target_text("Target: (none)")
            return

        extinf = str(r.get("extinf_display", ""))
        self._set_target_text(
            f"Target: {self._active_target} | key={self._active_pl_key} | row={self._active_row_id} | {extinf}"
        )

        cands = list(r.get("candidates", []) or [])

        # In Resolved, show the current manual choice rather than visually
        # reverting to the report's original automatic candidate.  The table row
        # already contains the merged manual state; mirror that state here.
        current_manual = None
        in_memory = (self._selections_by_key.get(self._active_pl_key, {}) or {}).get(self._active_row_id)
        if in_memory and (
            self._view_mode != "RESOLVED"
            or self._selection_origin.get((self._active_pl_key, self._active_row_id)) == "RESOLVED"
        ):
            current_manual = in_memory
        if not current_manual and self._active_pl_key in self._persisted_progress_keys:
            current_manual = (self._load_selections_for_key(self._active_pl_key) or {}).get(self._active_row_id)
        if current_manual:
            current_manual = str(current_manual)
            cands = [current_manual] + [p for p in cands if str(p) != current_manual]

        if self._active_target == "FAILED" and not current_manual:
            self.lst_candidates.addItem(QListWidgetItem("(No candidates. Use Browse…)"))
            return

        if not cands:
            self.lst_candidates.addItem(QListWidgetItem("(No candidates parsed from Notes. Use Browse…)"))
            return

        for p in cands:
            item = QListWidgetItem(p)
            item.setData(Qt.UserRole, p)
            self.lst_candidates.addItem(item)
        self.lst_candidates.setCurrentRow(0)

    def _selected_visual_row(self, table: QTableWidget) -> int | None:
        sel = table.selectionModel()
        if not sel or not sel.hasSelection():
            return None
        idxs = sel.selectedRows()
        if not idxs:
            return None
        return idxs[0].row()

    def _selected_row_id(self, table: QTableWidget) -> tuple[str | None, str | None]:
        vis_row = self._selected_visual_row(table)
        if vis_row is None:
            return None, None

        item0 = table.item(vis_row, 1)
        if not item0:
            return None, None

        data = item0.data(Qt.UserRole)
        if not isinstance(data, dict):
            return None, None

        pl_key = data.get("pl_key")
        row_id = data.get("row_id")
        if pl_key is None or row_id is None:
            return None, None

        return str(pl_key), str(row_id)

    def _current_candidate(self) -> str:
        # Prefer the explicitly selected item.  currentItem() can remain on row 0
        # in some Qt focus transitions (for example after clicking Apply), which
        # made Ambiguous repairs silently use the first candidate.
        selected = self.lst_candidates.selectedItems()
        it = selected[0] if selected else self.lst_candidates.currentItem()
        if not it:
            return ""
        data_path = it.data(Qt.UserRole)
        txt = str(data_path if data_path else (it.text() or "")).strip()
        if txt.startswith("("):
            return ""
        return txt

    def on_cancel_task(self):
        if not self._busy or self.worker is None:
            return
        self.worker.cancel()
        self.btn_cancel.setEnabled(False)
        self.status_label.setText("Cancelling safely…")

    def _clear_interrupted_action(self, action: str | None) -> None:
        if action in ("REMOVE_ROOTS", "CLEAR_ROOTS"):
            self._pending_remove_paths = set()
        elif action == "SAVE":
            self._pending_save_keys = []
            self._pending_save_snapshots = []
        elif action == "REPAIR":
            interrupted = set(self._pending_repair_keys)
            self._session_repaired_keys.difference_update(interrupted)
            self._provisional_reset_keys.difference_update(interrupted)
            self._pending_repair_keys = set()
            self._reload_reports_cache()
            self._refresh_tables_from_mode()
        self._last_action = None

    def _run_task(self, func, **kwargs):
        if self._busy or (self.thread is not None and self.thread.isRunning()):
            QMessageBox.information(self, "Busy", "A task is already running. Please wait.")
            return

        self._last_progress_msg = ""
        self._set_busy(True, "Running...")
        self.progress.setValue(0)

        self.thread = QThread()
        self.worker = Worker(func, kwargs)
        self.worker.moveToThread(self.thread)

        self.thread.started.connect(self.worker.run)
        self.worker.progress.connect(self._on_progress)
        self.worker.finished.connect(self._on_finished)
        self.worker.failed.connect(self._on_failed)

        self.worker.finished.connect(self.thread.quit)
        self.worker.failed.connect(self.thread.quit)
        self.worker.finished.connect(self.worker.deleteLater)
        self.worker.failed.connect(self.worker.deleteLater)
        self.thread.finished.connect(self._on_thread_finished)
        self.thread.finished.connect(self.thread.deleteLater)

        self.thread.start()

    @Slot()
    def _on_thread_finished(self):
        self.worker = None
        self.thread = None
        self._set_busy(False)
        if self._close_when_finished:
            self._close_when_finished = False
            self._close_without_prompt = True
            self.close()

    @Slot(int, str)
    def _on_progress(self, pct: int, msg: str):
        self.progress.setValue(int(pct))
        if msg:
            self._last_progress_msg = msg
            self.status_label.setText(msg)
            m = re.search(r"(?:indexed so far|Indexed)\s*:\s*(\d+)", msg, re.IGNORECASE)
            if m:
                self.scan_count_label.setText(f"Indexed: {m.group(1)}")

    @Slot(object)
    def _on_finished(self, result):
        self.progress.setValue(100)

        if isinstance(result, TaskResult):
            self.status_label.setText(result.message or "Done")
            outs = result.outputs or {}
            if not result.ok:
                action = self._last_action
                cancelled = bool(outs.get("cancelled")) or "cancel" in (result.message or "").casefold()
                self._clear_interrupted_action(action)
                self.progress.setValue(0)
                self.status_label.setText("Cancelled" if cancelled else "Error")
                if not cancelled and not self._close_when_finished:
                    QMessageBox.critical(self, "Error", result.message or "The task failed.")
                return

            if getattr(self, "_last_action", None) in ("REMOVE_ROOTS", "CLEAR_ROOTS"):
                action = self._last_action
                pending = set(getattr(self, "_pending_remove_paths", set()) or set())
                self._last_action = None
                self._pending_remove_paths = set()
                if not result.ok:
                    title = "Clear failed" if action == "CLEAR_ROOTS" else "Remove failed"
                    QMessageBox.critical(self, title, result.message)
                    return

                if action == "CLEAR_ROOTS":
                    self.music_roots = []
                else:
                    self.music_roots = [
                        root
                        for root in self.music_roots
                        if str(root.get("path", "")) not in pending
                    ]
                    # Scope aliases whose index owner was removed become Pending
                    # roots; nested aliases attach to the shallowest survivor.
                    self._reconcile_scope_roots(self._indexed_root_paths())
                self._save_music_roots()
                self._refresh_music_roots_ui()
                total = int(outs.get("total_indexed", 0) or 0)
                removed = int(outs.get("removed_indexed", 0) or 0)
                reassigned = int(outs.get("reassigned_indexed", 0) or 0)
                seconds = float(outs.get("remove_seconds", 0) or 0)
                self.scan_count_label.setText(f"Indexed: {total}")
                if action == "CLEAR_ROOTS":
                    message = (
                        f"Cleared all music folders and removed {removed} indexed track(s) "
                        f"in {seconds:.2f}s"
                    )
                else:
                    if reassigned:
                        message = (
                            f"Removed {len(pending)} folder(s); kept {reassigned} indexed "
                            f"track(s) for remaining child folders and removed {removed}; "
                            f"total library: {total}; completed in {seconds:.2f}s"
                        )
                    else:
                        message = (
                            f"Removed {len(pending)} folder(s) and {removed} indexed track(s); "
                            f"total library: {total}; completed in {seconds:.2f}s"
                        )
                self.status_label.setText(message)
                return

            # Scan result: persist only roots that actually contributed tracks.
            if "root_results" in outs:
                self._last_action = None
                results = outs.get("root_results", []) or []
                full_results = [
                    result_row
                    for result_row in results
                    if (
                        result_row.get("root")
                        and normalized_root_path(result_row.get("root", ""))
                        == normalized_root_path(
                            result_row.get("scan_path") or result_row.get("root", "")
                        )
                    )
                ]
                scanned_paths = {
                    str(Path(r.get("root", "")))
                    for r in full_results
                    if r.get("root")
                }
                successful = {
                    str(Path(r.get("root", "")))
                    for r in full_results
                    if r.get("root") and int(r.get("indexed", 0) or 0) > 0
                }
                successful_targets = sum(
                    1 for r in results if int(r.get("indexed", 0) or 0) > 0
                )

                kept: list[dict] = []
                for entry in self.music_roots:
                    key = str(Path(entry.get("path", "")))
                    if key in scanned_paths:
                        # Keep the folder visible even if this scan found zero
                        # tracks. It may be temporarily unavailable, empty, or
                        # contain files the scanner cannot currently read.
                        entry["imported"] = key in successful
                        entry["index_missing"] = False
                    kept.append(entry)

                self.music_roots = kept
                # A successful parent scan takes ownership of covered child
                # records. Keep those child paths as scope aliases, not new scans.
                self._reconcile_scope_roots(self._indexed_root_paths())
                self._save_music_roots()
                self._refresh_music_roots_ui()

                rejected = len(scanned_paths - successful)
                scan_indexed = int(outs.get("scan_indexed", 0) or 0)
                total_indexed = int(outs.get("total_indexed", scan_indexed) or 0)
                preserved_indexed = int(outs.get("preserved_indexed", 0) or 0)
                deduplicated_paths = int(outs.get("deduplicated_paths", 0) or 0)
                self.scan_count_label.setText(f"Indexed: {total_indexed}")
                msg = (
                    f"Scan complete. Added/updated folders: {successful_targets}; "
                    f"tracks scanned: {scan_indexed}; total library: {total_indexed}"
                )
                if preserved_indexed:
                    msg += f"; retained from other indexed roots: {preserved_indexed}"
                if deduplicated_paths:
                    msg += f"; duplicate paths removed: {deduplicated_paths}"
                if rejected:
                    msg += f"; empty/unreadable roots kept as Pending: {rejected}"
                self.status_label.setText(msg)
                if not self._close_when_finished:
                    QMessageBox.information(self, "Scan Complete", msg)
                return

            # Repair result
            if "ambiguous" in outs or "failed" in outs:
                self._last_action = None
                self._pending_repair_keys = set()
                # After repair, reports on disk changed -> reload cache then refresh current view
                self._reload_reports_cache()
                self._refresh_tables_from_mode()

                summaries = outs.get("summaries", []) or []
                if summaries:
                    try:
                        total_amb = sum(int(s.get("ambiguous", 0)) for s in summaries)
                        total_fail = sum(int(s.get("failed", 0)) for s in summaries)
                        total_rep = sum(int(s.get("repaired", 0)) for s in summaries)
                        total_kept = sum(int(s.get("kept", 0)) for s in summaries)
                    except Exception:
                        total_amb = total_fail = total_rep = total_kept = 0

                    if not self._close_when_finished:
                        QMessageBox.information(
                            self,
                            "Repair Complete",
                            f"Kept: {total_kept}\n"
                            f"Repaired: {total_rep}\n"
                            f"Ambiguous: {total_amb}\n"
                            f"Failed: {total_fail}\n\n"
                            f"Reports: {self.reports_path}",
                        )
                return

            # Save result
            if "done" in outs:
                if getattr(self, "_last_action", None) == "SAVE":
                    try:
                        for snapshot in getattr(self, "_pending_save_snapshots", []) or []:
                            pl_key = snapshot["source_key"]
                            self._commit_saved_snapshot(snapshot)
                            self._saved_keys.add(pl_key)
                            self._persisted_progress_keys.add(pl_key)
                            self._provisional_reset_keys.discard(pl_key)
                            self._dirty_selection_ids = {
                                key for key in self._dirty_selection_ids if key[0] != pl_key
                            }
                            self._selection_origin = {
                                key: origin
                                for key, origin in self._selection_origin.items()
                                if key[0] != pl_key
                            }
                    except Exception as exc:
                        self._pending_save_keys = []
                        self._pending_save_snapshots = []
                        self._last_action = None
                        self.progress.setValue(0)
                        self.status_label.setText("Repair history was not saved")
                        if not self._close_when_finished:
                            QMessageBox.critical(
                                self,
                                "Repair history not saved",
                                "The fixed playlist was written, but its repair history could "
                                f"not be saved completely.\n\n{exc}",
                            )
                        return
                    self._pending_save_keys = []
                    self._pending_save_snapshots = []
                    self._last_action = None

                done = outs.get("done", [])
                first = done[0].get("out_m3u", "") if done else ""
                if not self._close_when_finished:
                    QMessageBox.information(
                        self,
                        "Save Complete",
                        f"{result.message}\n\nExample output:\n{first}",
                    )

                # After save, if user is in Unresolved view, they likely want remaining list updated.
                # Reload reports cache + refresh tables (respects current view mode).
                self._reload_reports_cache()
                self._refresh_tables_from_mode()
                return

        self._last_action = None
        self.status_label.setText("Done")

    @Slot(str)
    def _on_failed(self, err: str):
        action = self._last_action
        self._clear_interrupted_action(action)
        self.progress.setValue(0)
        self.status_label.setText("Error")
        if not self._close_when_finished:
            QMessageBox.critical(self, "Error", err)

    def closeEvent(self, event):
        if self._close_without_prompt:
            event.accept()
            return

        thread_running = self.thread is not None and self.thread.isRunning()
        if thread_running:
            detail = ""
            if self._has_unsaved_work():
                detail = "\n\nUnsaved repair changes will also be discarded."
            answer = QMessageBox.question(
                self,
                "Cancel task and close?",
                "A background task is still running. Cancel it safely and close "
                "after it has stopped?"
                + detail,
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if answer != QMessageBox.Yes:
                event.ignore()
                return
            self._close_when_finished = True
            if self.worker is not None:
                self.worker.cancel()
            self.btn_cancel.setEnabled(False)
            self.status_label.setText("Cancelling safely before closing…")
            event.ignore()
            return

        if not self._confirm_discard_unsaved("Closing Playlist Fixer"):
            event.ignore()
            return
        event.accept()

    def _fill_table(self, table: QTableWidget, rows: list[dict]):
        """Fill the repair table efficiently, including large Resolved views."""
        sorting = table.isSortingEnabled()
        old_updates = table.updatesEnabled()
        old_blocked = table.blockSignals(True)
        table.setSortingEnabled(False)
        table.setUpdatesEnabled(False)
        table.setRowCount(len(rows))

        try:
            for vis_row, r in enumerate(rows):
                playlist = str(r.get("playlist", ""))
                pl_key = str(r.get("pl_key", ""))
                row_id = str(r.get("row_index", "")).strip()
                bucket = str(r.get("_bucket", ""))
                extinf = str(r.get("extinf_display", ""))
                orig = str(r.get("original_path", ""))
                notes = str(r.get("notes", ""))

                dirty = (pl_key, row_id) in self._dirty_selection_ids
                if dirty:
                    symbol = "✓"
                elif self._view_mode == "RESOLVED":
                    symbol = ""
                elif bucket == "AMBIGUOUS":
                    symbol = "△"
                elif bucket == "FAILED":
                    symbol = "✕"
                else:
                    symbol = ""

                status_item = QTableWidgetItem(symbol)
                status_item.setTextAlignment(Qt.AlignCenter)
                status_item.setData(Qt.UserRole, {"bucket": bucket})
                table.setItem(vis_row, 0, status_item)

                it1 = QTableWidgetItem(playlist)
                it1.setData(Qt.UserRole, {"pl_key": pl_key, "row_id": row_id})
                table.setItem(vis_row, 1, it1)
                table.setItem(vis_row, 2, QTableWidgetItem(extinf))
                table.setItem(vis_row, 3, QTableWidgetItem(orig))
                table.setItem(vis_row, 4, QTableWidgetItem(notes))
        finally:
            table.blockSignals(old_blocked)
            table.setUpdatesEnabled(old_updates)
            table.setSortingEnabled(sorting)
            table.viewport().update()

    def _norm(self, s: str) -> str:
        s = (s or "").strip().lower()
        # 讓 "A - B" / "A_B" / 空白差異比較不影響
        s = s.replace("\u3000", " ")
        return " ".join(s.split())

    def _row_matches_query(self, r: dict, q: str) -> bool:
        """
        q: already normalized lower
        Match against:
        - EXTINF display (song title)
        - original path (full + basename)
        - notes (includes [SELECTED]/[RESCUED]/[AUTO]/[MANUAL] paths)
        - candidates list (full + basename)
        - playlist name (optional but handy)
        """
        if not q:
            return True

        def hit(text: str) -> bool:
            t = self._norm(text)
            return bool(t) and (q in t)

        # Playlist filename
        if hit(Path(str(r.get("playlist", ""))).name):
            return True

        # EXTINF
        if hit(str(r.get("extinf_display", ""))):
            return True

        # Original path
        orig = str(r.get("original_path", "") or "")
        if orig and (hit(orig) or hit(Path(orig).name)):
            return True

        # Notes
        notes = str(r.get("notes", "") or "")
        if notes and hit(notes):
            return True

        # Candidates
        cands = r.get("candidates", []) or []
        for p in cands:
            ps = str(p or "")
            if ps and (hit(ps) or hit(Path(ps).name)):
                return True

        return False

    def _apply_search_filter(self) -> None:
        """
        Apply UI-only filtering using self._ambiguous_rows_all / self._failed_rows_all as masters.
        IMPORTANT:
        - Do NOT mutate *_rows_all here (they must remain unfiltered masters).
        - Do NOT call itself (no recursion).
        """
        q = ""
        if hasattr(self, "edt_search") and self.edt_search is not None:
            q = self._norm(self.edt_search.text())

        self._ambiguous_rows = [r for r in (self._ambiguous_rows_all or []) if self._row_matches_query(r, q)]
        self._failed_rows = [r for r in (self._failed_rows_all or []) if self._row_matches_query(r, q)]

        for r in self._ambiguous_rows:
            r.setdefault("_bucket", "AMBIGUOUS")
        for r in self._failed_rows:
            r["_bucket"] = "FAILED"

        self._rebuild_row_maps()
        if self._view_mode == "RESOLVED":
            # Resolved is an audit view, so preserve the original playlist row
            # order instead of grouping manual Ambiguous/Failed origins ahead
            # of automatic results.
            visible_rows = sorted(
                self._ambiguous_rows + self._failed_rows,
                key=lambda r: (
                    int(r.get("_playlist_order", 0)),
                    int(r.get("_row_order", 10**12)),
                ),
            )
        else:
            # Unresolved intentionally shows Ambiguous before Failed.
            visible_rows = self._ambiguous_rows + self._failed_rows
        self._fill_table(self.tbl_repair, visible_rows)

        # optional: clear selection after filtering to avoid stale row_id
        self.tbl_repair.clearSelection()
        self.lst_candidates.clear()
        self._active_target = None
        self._active_pl_key = None
        self._active_row_id = None
        self._set_target_text("Target: (none)")

    def on_search_changed(self, _text: str) -> None:
        self._apply_search_filter()
